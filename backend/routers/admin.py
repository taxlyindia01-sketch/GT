# routers/admin.py — Taxly super-admin endpoints
# Admin credentials: username=Taxly, password=@Gsf025@

from datetime import datetime, timezone
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from database import get_db
from models import Tenant, User, Invoice, Customer, ApprovalStatus
from utils.auth import (
    verify_password, hash_password, create_access_token, require_taxly_admin
)
from config import settings

router = APIRouter()


# ── Admin Login ───────────────────────────────────────────────

class AdminLoginRequest(BaseModel):
    username: str
    password: str

@router.post("/login")
async def admin_login(body: AdminLoginRequest):
    """
    Taxly super-admin login.
    Credentials: username='Taxly', password='@Gsf025@'
    """
    if body.username != settings.TAXLY_ADMIN_USERNAME:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not verify_password(body.password, settings.ADMIN_PASSWORD_HASH):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_access_token({
        "sub":             "taxly-admin",
        "is_taxly_admin":  True,
        "role":            "super_admin",
    })
    return {"access_token": token, "token_type": "bearer", "username": "Taxly"}


# ── Tenant Management ─────────────────────────────────────────

@router.get("/tenants")
async def list_tenants(
    _:  dict          = Depends(require_taxly_admin),
    db: AsyncSession  = Depends(get_db),
):
    """List all tenants with invoice and user counts."""
    result  = await db.execute(select(Tenant).order_by(Tenant.id))
    tenants = result.scalars().all()

    rows = []
    for t in tenants:
        inv_count  = await db.scalar(select(func.count()).where(Invoice.tenant_id  == t.id)) or 0
        user_count = await db.scalar(select(func.count()).where(User.tenant_id     == t.id)) or 0
        rows.append({
            "id":           t.id,
            "company_name": t.company_name,
            "plan":         t.plan.value,
            "is_active":    t.is_active,
            "user_count":   user_count,
            "invoice_count":inv_count,
            "created_at":   t.created_at.isoformat(),
        })
    return rows


@router.post("/tenants", status_code=201)
async def create_tenant(
    body: dict,
    _:   dict         = Depends(require_taxly_admin),
    db:  AsyncSession = Depends(get_db),
):
    """Create a new tenant with initial admin user."""
    tenant = Tenant(company_name=body["company_name"], plan=body.get("plan", "demo"))
    db.add(tenant)
    await db.flush()

    admin_user = User(
        tenant_id=tenant.id,
        username=body["admin_username"],
        mobile=body["admin_mobile"],
        password_hash=hash_password(body["password"]),
        role="admin",
        is_active=True,
        approval_status=ApprovalStatus.approved,
    )
    db.add(admin_user)
    await db.commit()
    return {"tenant_id": tenant.id, "message": "Tenant created successfully"}


@router.patch("/tenants/{tenant_id}/toggle")
async def toggle_tenant(
    tenant_id: int,
    _:   dict         = Depends(require_taxly_admin),
    db:  AsyncSession = Depends(get_db),
):
    """Enable or disable a tenant."""
    tenant = await db.get(Tenant, tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    tenant.is_active = not tenant.is_active
    await db.commit()
    return {"tenant_id": tenant_id, "is_active": tenant.is_active}


@router.patch("/tenants/{tenant_id}/reset-password")
async def reset_tenant_password(
    tenant_id: int,
    body:      dict,
    _:   dict         = Depends(require_taxly_admin),
    db:  AsyncSession = Depends(get_db),
):
    """Reset the admin password for a tenant."""
    result = await db.execute(
        select(User).where(User.tenant_id == tenant_id, User.role == "admin")
    )
    admin = result.scalars().first()
    if not admin:
        raise HTTPException(status_code=404, detail="Admin user not found for this tenant")
    admin.password_hash = hash_password(body["new_password"])
    await db.commit()
    return {"message": f"Password reset for {admin.username}"}


# ── User Management ───────────────────────────────────────────

@router.get("/users")
async def list_all_users(
    _:  dict         = Depends(require_taxly_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all users across all tenants."""
    result  = await db.execute(select(User).order_by(User.tenant_id, User.id))
    users   = result.scalars().all()

    tenant_cache = {}
    rows = []
    for u in users:
        if u.tenant_id not in tenant_cache:
            t = await db.get(Tenant, u.tenant_id)
            tenant_cache[u.tenant_id] = t.company_name if t else "Unknown"
        rows.append({
            "id":              u.id,
            "username":        u.username,
            "mobile":          u.mobile,
            "email":           u.email,
            "role":            u.role.value,
            "tenant":          tenant_cache[u.tenant_id],
            "tenant_id":       u.tenant_id,
            "is_active":       u.is_active,
            "auth_provider":   u.auth_provider.value,
            "approval_status": u.approval_status.value,
        })
    return rows


# ── Google Signup Requests ────────────────────────────────────

@router.get("/google-requests")
async def list_google_requests(
    _:  dict         = Depends(require_taxly_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    List all users who signed up via Google.
    Includes trial users (10-day free) and pending approvals (trial expired).
    """
    result = await db.execute(
        select(User).where(
            User.auth_provider == "google",
            User.approval_status.in_(["trial", "pending", "approved", "rejected"])
        ).order_by(User.created_at.desc())
    )
    users = result.scalars().all()

    now = datetime.now(timezone.utc)
    rows = []
    for u in users:
        days_left = 0
        if u.trial_expires_at and u.approval_status.value == "trial":
            delta     = u.trial_expires_at - now
            days_left = max(0, delta.days)
            if days_left == 0:
                u.approval_status = ApprovalStatus.pending

        rows.append({
            "id":              u.id,
            "name":            u.username,
            "email":           u.email,
            "company":         u.company_name,
            "mobile":          u.mobile,
            "status":          u.approval_status.value,
            "trial_expires_at":u.trial_expires_at.isoformat() if u.trial_expires_at else None,
            "days_left":       days_left,
            "signed_up":       u.created_at.date().isoformat(),
        })

    await db.commit()   # commit any pending→trial changes
    return rows


@router.patch("/google-requests/{user_id}/approve")
async def approve_google_user(
    user_id:   int,
    body:      dict = {},
    _:   dict         = Depends(require_taxly_admin),
    db:  AsyncSession = Depends(get_db),
):
    """
    Approve a Google signup user.
    Sets status → approved, unlocks full CRM access.
    In production: also sends approval email via SMTP.
    """
    user = await db.get(User, user_id)
    if not user or user.auth_provider.value != "google":
        raise HTTPException(status_code=404, detail="Google user not found")

    user.approval_status  = ApprovalStatus.approved
    user.trial_expires_at = None   # clear trial expiry on approval
    user.role             = body.get("role", user.role.value)
    await db.commit()

    # TODO: send_approval_email(user.email, user.username)
    return {"message": f"{user.username} approved successfully. They can now log in."}


@router.patch("/google-requests/{user_id}/reject")
async def reject_google_user(
    user_id: int,
    _:   dict         = Depends(require_taxly_admin),
    db:  AsyncSession = Depends(get_db),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.approval_status = ApprovalStatus.rejected
    await db.commit()
    return {"message": f"{user.username} rejected."}


# ── Tenant Backups ────────────────────────────────────────────

@router.get("/backups")
async def list_backups(
    _:  dict         = Depends(require_taxly_admin),
    db: AsyncSession = Depends(get_db),
):
    """List backup info for all tenants."""
    result  = await db.execute(select(Tenant).order_by(Tenant.id))
    tenants = result.scalars().all()

    rows = []
    for t in tenants:
        inv_count  = await db.scalar(select(func.count()).where(Invoice.tenant_id  == t.id)) or 0
        cust_count = await db.scalar(select(func.count()).where(Customer.tenant_id == t.id)) or 0
        rows.append({
            "tenant_id":    t.id,
            "company_name": t.company_name,
            "is_active":    t.is_active,
            "invoice_count":inv_count,
            "customer_count":cust_count,
        })
    return rows


@router.get("/backups/{tenant_id}/download")
async def download_tenant_backup(
    tenant_id: int,
    _:   dict         = Depends(require_taxly_admin),
    db:  AsyncSession = Depends(get_db),
):
    """
    Download full Excel backup for a specific tenant.
    Generates a multi-sheet workbook: Invoices, Customers, Payments,
    Cash Register, Advances, Stock, Cancelled Invoices.
    Fix: previously returned a stub JSON — now streams a real Excel file.
    """
    from io import BytesIO
    from datetime import datetime as dt, date as date_type
    from fastapi.responses import StreamingResponse
    from sqlalchemy import select as sa_select
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from models import (
        Invoice, Customer, Payment, CashEntry, Advance, StockItem, StockTransaction
    )

    tenant = await db.get(Tenant, tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    # ── helpers ───────────────────────────────────────────────
    GOLD_FILL   = PatternFill("solid", fgColor="C8900A")
    HDR_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

    def add_sheet(wb, title, headers, rows):
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = GOLD_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)
        return ws

    tid = tenant_id
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1 — Invoices
    r = await db.execute(sa_select(Invoice).where(Invoice.tenant_id == tid, Invoice.status == "active").order_by(Invoice.invoice_date.desc()))
    invoices = r.scalars().all()
    add_sheet(wb, "Invoices",
        ["Invoice No","Date","Customer","Mobile","PAN","Mode","Subtotal","CGST","SGST","IGST","Grand Total","Paid","Outstanding","Status"],
        [[i.invoice_no, i.invoice_date.isoformat(), i.customer_name, i.customer_mobile,
          i.customer_pan or "", i.pay_mode.value, float(i.subtotal), float(i.cgst),
          float(i.sgst), float(i.igst), float(i.grand_total), float(i.amount_paid),
          float(i.outstanding), i.payment_status.value] for i in invoices])

    # Sheet 2 — Cancelled Invoices
    r2 = await db.execute(sa_select(Invoice).where(Invoice.tenant_id == tid, Invoice.status == "cancelled").order_by(Invoice.updated_at.desc()))
    cancelled = r2.scalars().all()
    add_sheet(wb, "Cancelled Invoices",
        ["Invoice No","Invoice Date","Cancelled At","Customer","Mobile","PAN","Mode","Grand Total","Notes"],
        [[i.invoice_no, i.invoice_date.isoformat(),
          i.updated_at.isoformat() if i.updated_at else "",
          i.customer_name, i.customer_mobile, i.customer_pan or "",
          i.pay_mode.value, float(i.grand_total), i.notes or ""] for i in cancelled])

    # Sheet 3 — Customers
    r = await db.execute(sa_select(Customer).where(Customer.tenant_id == tid).order_by(Customer.name))
    customers = r.scalars().all()
    add_sheet(wb, "Customers",
        ["Mobile","Name","PAN","State","GSTIN","Address","Cash Receipts FY","SFT Flagged"],
        [[c.mobile, c.name, c.pan or "", c.state, c.gstin or "",
          c.address or "", float(c.cash_receipts_fy), "Yes" if c.sft_flagged else "No"]
         for c in customers])

    # Sheet 4 — Payments
    inv_map = {i.id: i.invoice_no for i in invoices}
    r = await db.execute(sa_select(Payment).where(Payment.tenant_id == tid).order_by(Payment.payment_date.desc()))
    payments = r.scalars().all()
    add_sheet(wb, "Payments",
        ["ID","Invoice No","Customer Mobile","Amount","Date","Mode","Reference","Notes"],
        [[p.id, inv_map.get(p.invoice_id, ""), p.customer_mobile, float(p.amount),
          p.payment_date.isoformat(), p.pay_mode.value,
          p.reference_no or "", p.notes or ""] for p in payments])

    # Sheet 5 — Cash Register
    r = await db.execute(sa_select(CashEntry).where(CashEntry.tenant_id == tid).order_by(CashEntry.entry_date.desc()))
    cash = r.scalars().all()
    add_sheet(wb, "Cash Register",
        ["Date","Type","Description","Amount","Bank Reference"],
        [[e.entry_date.isoformat(), e.entry_type.value, e.description,
          float(e.amount), e.bank_reference or ""] for e in cash])

    # Sheet 6 — Advances
    r = await db.execute(sa_select(Advance).where(Advance.tenant_id == tid))
    advances = r.scalars().all()
    add_sheet(wb, "Advances",
        ["ID","Customer Mobile","Amount","Remaining","Date","Mode","Notes"],
        [[a.id, a.customer_mobile, float(a.amount), float(a.remaining),
          a.advance_date.isoformat(), a.pay_mode.value, a.notes or ""]
         for a in advances])

    # Sheet 7 — Stock
    r = await db.execute(sa_select(StockItem).where(StockItem.tenant_id == tid))
    stocks = r.scalars().all()
    add_sheet(wb, "Stock",
        ["ID","Category","Purity","Description","Unit","Qty on Hand"],
        [[s.id, s.category.value, s.purity or "", s.description,
          s.unit.value, float(s.qty_on_hand)] for s in stocks])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = tenant.company_name.replace(" ", "_").replace("/", "-")
    fname = f"backup_{safe}_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
