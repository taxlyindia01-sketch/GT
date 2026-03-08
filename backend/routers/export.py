# routers/export.py
# Changes vs v4 original:
#  Issue 7/10 — /payments-excel: uses date-filtered payments data (was broken)
#  Issue 8    — /advances-excel: new endpoint for Advances page download
#  P11        — TCS sheet replaced with Section 269ST sheet in full backup

from io import BytesIO
from datetime import date, datetime
from decimal import Decimal
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models import (
    Invoice, InvoiceItem, Customer, Payment,
    CashEntry, Advance, StockItem, StockTransaction,
    Supplier, SupplierInvoice, SupplierInvoiceItem, SupplierPayment, SupplierAdvance,
)
from utils.auth import get_current_user_payload
from utils.business import current_fy, SFT_THRESHOLD

router = APIRouter()

# ── Styling ───────────────────────────────────────────────────

GOLD_FILL   = PatternFill("solid", fgColor="C8900A")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="Calibri", size=10)
BORDER      = Border(
    bottom=Side(style="thin", color="E8B840"),
    top=Side(style="thin",    color="E8B840"),
)


def style_header_row(ws, row_num: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = GOLD_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_col_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


_LIGHT_ALT  = PatternFill("solid", fgColor="F5F7FA")
_LIGHT_EVEN = PatternFill("solid", fgColor="FFFFFF")
_BODY_FONT  = Font(name="Calibri", size=10, color="333333")

def add_sheet(wb, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for i, row in enumerate(rows):
        ws.append(row)
        fill = _LIGHT_ALT if i % 2 == 0 else _LIGHT_EVEN
        for col in range(1, len(row) + 1):
            cell = ws.cell(i + 2, col)
            cell.fill = fill
            if cell.font == openpyxl.styles.DEFAULT_FONT:
                cell.font = _BODY_FONT
    auto_col_width(ws)
    ws.sheet_view.showGridLines = False

async def add_account_sheet(
    wb, db, tenant_id: int,
    from_date=None, to_date=None,
    invoices_cache=None,
) -> None:
    """Add Account Register sheet to workbook (category-wise invoice breakdown)."""
    from decimal import Decimal
    if invoices_cache is not None:
        invoices = invoices_cache
    else:
        stmt = select(Invoice).where(Invoice.tenant_id == tenant_id, Invoice.status == "active")
        if from_date: stmt = stmt.where(Invoice.invoice_date >= from_date)
        if to_date:   stmt = stmt.where(Invoice.invoice_date <= to_date)
        result   = await db.execute(stmt.order_by(Invoice.invoice_date.desc()))
        invoices = result.scalars().all()

    headers = [
        "Invoice Date", "Invoice No", "Customer Name", "Customer Mobile",
        "Gold (₹)", "Silver (₹)", "Diamond (₹)", "Polish Charges (₹)",
        "Making Charges (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Grand Total (₹)"
    ]
    ws = wb.create_sheet(title="Account Register")
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    tot = {k: Decimal("0") for k in ["gold","silver","diamond","polish","making","cgst","sgst","igst","grand"]}

    for inv in invoices:
        items_result = await db.execute(select(InvoiceItem).where(InvoiceItem.invoice_id == inv.id))
        items = items_result.scalars().all()
        gold_amt = silver_amt = diamond_amt = polish_amt = making_total = Decimal("0")
        for item in items:
            cat = item.category.value
            item_base = item.amount - item.making_charges
            making_total += item.making_charges
            if cat == "Gold":             gold_amt    += item_base
            elif cat == "Silver":         silver_amt  += item_base
            elif cat == "Diamond":        diamond_amt += item_base
            elif cat == "Polish Charges": polish_amt  += item_base

        tot["gold"]    += gold_amt;    tot["silver"]  += silver_amt
        tot["diamond"] += diamond_amt; tot["polish"]  += polish_amt
        tot["making"]  += making_total; tot["cgst"]   += inv.cgst
        tot["sgst"]    += inv.sgst;    tot["igst"]    += inv.igst
        tot["grand"]   += inv.grand_total

        ws.append([
            inv.invoice_date.isoformat(), inv.invoice_no,
            inv.customer_name, inv.customer_mobile,
            float(gold_amt), float(silver_amt), float(diamond_amt), float(polish_amt),
            float(making_total),
            float(inv.cgst), float(inv.sgst), float(inv.igst), float(inv.grand_total),
        ])

    total_row_vals = [
        "TOTAL", "", "", "",
        float(tot["gold"]), float(tot["silver"]), float(tot["diamond"]), float(tot["polish"]),
        float(tot["making"]), float(tot["cgst"]), float(tot["sgst"]), float(tot["igst"]), float(tot["grand"]),
    ]
    ws.append(total_row_vals)
    tr = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=tr, column=col)
        cell.font    = Font(name="Calibri", bold=True, size=10)
        cell.fill    = PatternFill("solid", fgColor="FFF0CC")
    auto_col_width(ws)


async def add_dashboard_sheet(
    wb, db, tenant_id: int,
    from_date=None, to_date=None,
    invoices_cache=None,
) -> None:
    """Add Dashboard summary sheet to workbook."""
    from decimal import Decimal
    from datetime import date as date_type

    if invoices_cache is not None:
        invoices = invoices_cache
    else:
        stmt = select(Invoice).where(Invoice.tenant_id == tenant_id, Invoice.status == "active")
        if from_date: stmt = stmt.where(Invoice.invoice_date >= from_date)
        if to_date:   stmt = stmt.where(Invoice.invoice_date <= to_date)
        result   = await db.execute(stmt.order_by(Invoice.invoice_date.desc()))
        invoices = result.scalars().all()

    ws = wb.create_sheet(title="Dashboard")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    def kv(label, value):
        r = ws.max_row + 1
        ws.cell(r, 1, label).font = Font(name="Calibri", bold=True, size=10)
        cell = ws.cell(r, 2, value)
        cell.font = Font(name="Calibri", size=10)
        return cell

    def section(title):
        ws.append([])
        r = ws.max_row + 1
        c = ws.cell(r, 1, title)
        c.font = Font(name="Calibri", bold=True, size=12, color="C8900A")
        c.fill = PatternFill("solid", fgColor="FFF8E1")
        ws.cell(r, 2).fill = PatternFill("solid", fgColor="FFF8E1")

    today_dt = date_type.today()
    section("📊 Report Summary")
    kv("Generated On", today_dt.isoformat())
    kv("Period From",  from_date.isoformat() if from_date else "All time")
    kv("Period To",    to_date.isoformat()   if to_date   else today_dt.isoformat())
    kv("Total Invoices", len(invoices))

    # Financial summary
    total_subtotal  = sum(float(i.subtotal)    for i in invoices)
    total_cgst      = sum(float(i.cgst)        for i in invoices)
    total_sgst      = sum(float(i.sgst)        for i in invoices)
    total_igst      = sum(float(i.igst)        for i in invoices)
    total_gst       = total_cgst + total_sgst + total_igst
    total_grand     = sum(float(i.grand_total) for i in invoices)
    total_collected = sum(float(i.amount_paid) for i in invoices)
    total_outstanding = sum(float(i.outstanding) for i in invoices)

    section("💰 Financial Summary")
    kv("Total Subtotal (Taxable Value)",    f"Rs. {total_subtotal:,.2f}")
    kv("Total CGST",                        f"Rs. {total_cgst:,.2f}")
    kv("Total SGST",                        f"Rs. {total_sgst:,.2f}")
    kv("Total IGST",                        f"Rs. {total_igst:,.2f}")
    kv("Total GST",                         f"Rs. {total_gst:,.2f}")
    kv("Grand Total (All Invoices)",        f"Rs. {total_grand:,.2f}")
    kv("Total Collected (Payments)",        f"Rs. {total_collected:,.2f}")
    kv("Total Outstanding",                 f"Rs. {total_outstanding:,.2f}")

    # Payment mode breakdown
    from collections import Counter
    mode_totals: dict[str, float] = {}
    for inv in invoices:
        m = inv.pay_mode.value
        mode_totals[m] = mode_totals.get(m, 0) + float(inv.grand_total)

    section("💳 Sales by Payment Mode")
    for mode, amt in sorted(mode_totals.items(), key=lambda x: -x[1]):
        kv(mode, f"Rs. {amt:,.2f}")

    # Category breakdown (needs items)
    cat_totals: dict[str, float] = {"Gold": 0, "Silver": 0, "Diamond": 0, "Polish Charges": 0, "Making Charges": 0}
    for inv in invoices:
        items_result = await db.execute(select(InvoiceItem).where(InvoiceItem.invoice_id == inv.id))
        for item in items_result.scalars():
            cat = item.category.value
            base = float(item.amount - item.making_charges)
            if cat in cat_totals: cat_totals[cat] += base
            cat_totals["Making Charges"] += float(item.making_charges)

    section("🏷️ Sales by Category")
    for cat, amt in cat_totals.items():
        if amt > 0: kv(cat, f"Rs. {amt:,.2f}")

    # 269ST violations
    violations = [inv for inv in invoices if inv.pay_mode.value == "Cash" and float(inv.grand_total) >= 200000]
    section("⚠️ Section 269ST Compliance")
    kv("Threshold",                   "Rs. 2,00,000 per transaction")
    kv("Cash Invoices ≥ ₹2L (Violations)", len(violations))
    kv("Total Cash Violation Amount", f"Rs. {sum(float(i.grand_total) for i in violations):,.2f}")

    # Outstanding summary
    unpaid  = [i for i in invoices if i.payment_status.value == "unpaid"]
    partial = [i for i in invoices if i.payment_status.value == "partial"]
    paid    = [i for i in invoices if i.payment_status.value == "paid"]

    section("📋 Invoice Status")
    kv("Paid",    len(paid))
    kv("Partial", len(partial))
    kv("Unpaid",  len(unpaid))
    kv("Outstanding (Unpaid + Partial)", f"Rs. {total_outstanding:,.2f}")


def _stream_workbook(wb, filename: str) -> StreamingResponse:
    # Disable grid lines on every sheet before saving
    for ws_name in wb.sheetnames:
        wb[ws_name].sheet_view.showGridLines = False
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Full Backup (15 sheets) ───────────────────────────────────

@router.get("/excel")
async def export_full_backup(
    tenant_id: Optional[int] = Query(None),
    payload:   dict          = Depends(get_current_user_payload),
    db:        AsyncSession  = Depends(get_db),
):
    """
    Full Excel backup — 15 sheets.
    P11: TCS sheet replaced with Section 269ST violations sheet.
    """
    tid = tenant_id or payload["tenant_id"]
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)

    fy_start, fy_end = current_fy()

    # ── Sheet 1: Invoices ─────────────────────────────────────
    result   = await db.execute(select(Invoice).where(Invoice.tenant_id == tid).order_by(Invoice.invoice_date.desc()))
    invoices = result.scalars().all()

    add_sheet(wb, "Invoices", [
        "Invoice No", "Date", "Customer Name", "Mobile", "PAN", "Pay Mode",
        "Subtotal", "CGST", "SGST", "IGST", "Grand Total", "Paid", "Outstanding", "Status"
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name, inv.customer_mobile,
         inv.customer_pan or "", inv.pay_mode.value, float(inv.subtotal),
         float(inv.cgst), float(inv.sgst), float(inv.igst),
         float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding), inv.status.value]
        for inv in invoices
    ])

    # ── Sheet 2: Invoice Items ────────────────────────────────
    result = await db.execute(select(InvoiceItem).where(InvoiceItem.tenant_id == tid))
    items  = result.scalars().all()
    inv_no_map = {inv.id: inv.invoice_no for inv in invoices}

    add_sheet(wb, "Invoice_Items", [
        "Invoice No", "Category", "Purity", "Description", "HSN", "Qty", "Unit", "Rate", "Making", "Amount"
    ], [
        [inv_no_map.get(i.invoice_id, ""), i.category.value, i.purity or "", i.description,
         i.hsn_code, float(i.qty), i.unit.value, float(i.rate), float(i.making_charges), float(i.amount)]
        for i in items
    ])

    # ── Sheet 3: Payments ─────────────────────────────────────
    result   = await db.execute(select(Payment).where(Payment.tenant_id == tid).order_by(Payment.payment_date.desc()))
    payments = result.scalars().all()

    add_sheet(wb, "Payments", [
        "ID", "Invoice No", "Customer Name", "Customer Mobile", "Amount", "Date", "Mode", "Reference"
    ], [
        [p.id, inv_no_map.get(p.invoice_id, ""),
         getattr(p, "customer_name", "") or "",
         p.customer_mobile,
         float(p.amount), p.payment_date.isoformat(), p.pay_mode.value, p.reference_no or ""]
        for p in payments
    ])

    # ── Sheet 4: Customers ────────────────────────────────────
    result    = await db.execute(select(Customer).where(Customer.tenant_id == tid).order_by(Customer.name))
    customers = result.scalars().all()

    add_sheet(wb, "Customers", [
        "Mobile (PK)", "Name", "PAN", "State", "GSTIN", "Address", "Cash Receipts FY", "SFT Flagged"
    ], [
        [c.mobile, c.name, c.pan or "", c.state, c.gstin or "", c.address or "",
         float(c.cash_receipts_fy), "Yes" if c.sft_flagged else "No"]
        for c in customers
    ])

    # ── Sheet 5: Stock Items ──────────────────────────────────
    result = await db.execute(select(StockItem).where(StockItem.tenant_id == tid))
    stocks = result.scalars().all()

    add_sheet(wb, "Stock_Items", [
        "ID", "Category", "Purity", "Description", "Unit", "Qty on Hand"
    ], [
        [s.id, s.category.value, s.purity or "", s.description, s.unit.value, float(s.qty_on_hand)]
        for s in stocks
    ])

    # ── Sheet 6: Cash Register ────────────────────────────────
    result  = await db.execute(select(CashEntry).where(CashEntry.tenant_id == tid).order_by(CashEntry.entry_date.desc()))
    entries = result.scalars().all()

    add_sheet(wb, "Cash_Register", [
        "Date", "Type", "Description", "Amount", "Bank Reference"
    ], [
        [e.entry_date.isoformat(), e.entry_type.value, e.description,
         float(e.amount), e.bank_reference or ""]
        for e in entries
    ])

    # ── Sheet 7: Advances ─────────────────────────────────────
    result   = await db.execute(select(Advance).where(Advance.tenant_id == tid))
    advances = result.scalars().all()

    add_sheet(wb, "Advances", [
        "ID", "Customer Name", "Customer Mobile", "Amount", "Remaining", "Date", "Mode", "Notes"
    ], [
        [a.id,
         getattr(a, "customer_name", "") or "",
         a.customer_mobile,
         float(a.amount), float(a.remaining),
         a.advance_date.isoformat(), a.pay_mode.value, a.notes or ""]
        for a in advances
    ])

    # ── Sheet 8: Stock Transactions ───────────────────────────
    result = await db.execute(select(StockTransaction).where(StockTransaction.tenant_id == tid).order_by(StockTransaction.txn_date.desc()))
    txns   = result.scalars().all()

    add_sheet(wb, "Stock_Transactions", [
        "ID", "Stock Item ID", "Type", "Qty", "Purchase Rate", "Date", "Reason"
    ], [
        [t.id, t.stock_item_id, t.txn_type.value, float(t.qty),
         float(t.purchase_rate) if t.purchase_rate else "", t.txn_date.isoformat(), t.reason or ""]
        for t in txns
    ])

    # ── Report Sheet 9: Sales Register ───────────────────────
    fy_invoices = [inv for inv in invoices if fy_start <= inv.invoice_date <= fy_end and inv.status.value == "active"]

    add_sheet(wb, "Report_Sales", [
        "Invoice No", "Date", "Customer", "Mobile", "PAN", "HSN",
        "Subtotal", "CGST", "SGST", "IGST", "Grand Total", "Mode"
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name, inv.customer_mobile,
         inv.customer_pan or "", "7113", float(inv.subtotal),
         float(inv.cgst), float(inv.sgst), float(inv.igst),
         float(inv.grand_total), inv.pay_mode.value]
        for inv in fy_invoices
    ])

    # ── Report Sheet 10: Section 269ST (replaces TCS — P11) ───
    thresh_269st = Decimal("200000")
    viol_payments = [
        p for p in payments
        if p.pay_mode.value == "Cash" and p.amount >= thresh_269st
        and fy_start <= p.payment_date <= fy_end
    ]

    add_sheet(wb, "Report_Sec269ST", [
        "Date", "Invoice No", "Customer Name", "Mobile", "PAN", "Cash Amount", "Penalty Risk", "Reference"
    ], [])

    ws_269 = wb["Report_Sec269ST"]
    for p in viol_payments:
        cust  = await db.get(Customer, (p.customer_mobile, tid))
        cname = (getattr(p, "customer_name", None) or (cust.name if cust else "—"))
        cpan  = cust.pan if cust else ""
        ws_269.append([
            p.payment_date.isoformat(),
            inv_no_map.get(p.invoice_id, "—"),
            cname, p.customer_mobile, cpan or "MISSING",
            float(p.amount), float(p.amount), p.reference_no or "",
        ])
    auto_col_width(ws_269)

    # ── Report Sheet 11: SFT Register ────────────────────────
    sft_customers = [c for c in customers if c.sft_flagged]

    add_sheet(wb, "Report_SFT", [
        "Customer", "Mobile", "PAN", "Cash Receipts FY", "SFT Threshold", "PAN Missing"
    ], [
        [c.name, c.mobile, c.pan or "", float(c.cash_receipts_fy),
         float(SFT_THRESHOLD), "YES" if not c.pan else "No"]
        for c in sft_customers
    ])

    # ── Report Sheet 12: GSTR-1 ───────────────────────────────
    add_sheet(wb, "Report_GSTR1", [
        "Invoice No", "Date", "Customer", "GSTIN", "State", "HSN",
        "Taxable", "CGST%", "CGST", "SGST%", "SGST", "Total"
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name,
         inv.customer_gstin or "Unregistered", inv.customer_state or "",
         "7113", float(inv.subtotal),
         float(inv.gst_rate / 2), float(inv.cgst),
         float(inv.gst_rate / 2), float(inv.sgst),
         float(inv.grand_total)]
        for inv in fy_invoices
    ])

    # ── Report Sheet 13: Outstanding ──────────────────────────
    outstanding_invoices = [inv for inv in invoices if float(inv.outstanding) > 0]

    add_sheet(wb, "Report_Outstanding", [
        "Invoice No", "Date", "Customer", "Mobile", "Grand Total", "Paid", "Outstanding"
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name,
         inv.customer_mobile, float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding)]
        for inv in outstanding_invoices
    ])

    # ── Report Sheet 14: Cash Book ────────────────────────────
    fy_cash_entries = [e for e in entries if fy_start <= e.entry_date <= fy_end]

    add_sheet(wb, "Report_Cash_Book", [
        "Date", "Type", "Description", "Cash In", "Cash Out", "Bank In", "Balance"
    ], [
        [e.entry_date.isoformat(), e.entry_type.value, e.description,
         float(e.amount) if e.entry_type.value == "cash_in"                          else 0,
         float(e.amount) if e.entry_type.value in ("cash_out", "cash_to_bank")       else 0,
         float(e.amount) if e.entry_type.value == "bank_in"                          else 0,
         float(e.running_balance or 0)]
        for e in fy_cash_entries
    ])

    # ── Report Sheet 15: Payments Register ────────────────────
    add_sheet(wb, "Report_Payments", [
        "Date", "Invoice No", "Customer Name", "Mobile", "Amount", "Mode", "Reference"
    ], [
        [p.payment_date.isoformat(),
         inv_no_map.get(p.invoice_id, "—"),
         getattr(p, "customer_name", "") or "",
         p.customer_mobile,
         float(p.amount), p.pay_mode.value, p.reference_no or ""]
        for p in payments
    ])

    filename = f"goldtrader_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _stream_workbook(wb, filename)


# ── Payments Excel (standalone) ───────────────────────────────
# Issue 7/10 fix — now uses date-range filtered data

@router.get("/payments-excel")
async def export_payments_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Export payment register to Excel with optional date range."""
    tenant_id = payload["tenant_id"]
    stmt = (
        select(Payment)
        .where(Payment.tenant_id == tenant_id)
        .order_by(Payment.payment_date.desc())
    )
    if from_date:
        stmt = stmt.where(Payment.payment_date >= from_date)
    if to_date:
        stmt = stmt.where(Payment.payment_date <= to_date)

    result   = await db.execute(stmt)
    payments = result.scalars().all()

    # Build invoice_no map
    inv_result = await db.execute(select(Invoice).where(Invoice.tenant_id == tenant_id))
    inv_no_map = {inv.id: inv.invoice_no for inv in inv_result.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = ["Date", "Invoice No", "Customer Name", "Mobile", "Amount", "Mode", "Reference"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for p in payments:
        cname = getattr(p, "customer_name", None)
        if not cname:
            cust  = await db.get(Customer, (p.customer_mobile, tenant_id))
            cname = cust.name if cust else "—"
        ws.append([
            p.payment_date.isoformat(),
            inv_no_map.get(p.invoice_id, "—"),
            cname,
            p.customer_mobile,
            float(p.amount),
            p.pay_mode.value,
            p.reference_no or "",
        ])

    auto_col_width(ws)

    date_range = f"{from_date or 'all'}_{to_date or 'all'}"
    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"payment_register_{date_range}.xlsx")


# ── Advances Excel (new) ──────────────────────────────────────
# Issue 8 fix — new endpoint for Advances page "Download Excel" button

@router.get("/advances-excel")
async def export_advances_excel(
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    """Export advances register to Excel."""
    tenant_id = payload["tenant_id"]
    result    = await db.execute(
        select(Advance)
        .where(Advance.tenant_id == tenant_id)
        .order_by(Advance.advance_date.desc())
    )
    advances = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Advances"

    headers = ["Date", "Customer Name", "Mobile", "Amount", "Remaining", "Mode", "Notes"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for a in advances:
        cname = getattr(a, "customer_name", None)
        if not cname:
            cust  = await db.get(Customer, (a.customer_mobile, tenant_id))
            cname = cust.name if cust else "—"
        ws.append([
            a.advance_date.isoformat(),
            cname,
            a.customer_mobile,
            float(a.amount),
            float(a.remaining),
            a.pay_mode.value,
            a.notes or "",
        ])

    auto_col_width(ws)
    return _stream_workbook(wb, f"advances_register_{date.today().isoformat()}.xlsx")


# ── Sales Excel ───────────────────────────────────────────────

@router.get("/sales-excel")
async def export_sales_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Export sales register to Excel."""
    tenant_id = payload["tenant_id"]
    stmt = (
        select(Invoice)
        .where(Invoice.tenant_id == tenant_id, Invoice.status == "active")
        .order_by(Invoice.invoice_date.desc())
    )
    if from_date:
        stmt = stmt.where(Invoice.invoice_date >= from_date)
    if to_date:
        stmt = stmt.where(Invoice.invoice_date <= to_date)

    result   = await db.execute(stmt)
    invoices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    headers = [
        "Invoice No", "Date", "Customer", "Mobile", "PAN",
        "Pay Mode", "Subtotal", "CGST", "SGST", "IGST", "Grand Total", "Status"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for inv in invoices:
        ws.append([
            inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name,
            inv.customer_mobile, inv.customer_pan or "",
            inv.pay_mode.value, float(inv.subtotal),
            float(inv.cgst), float(inv.sgst), float(inv.igst),
            float(inv.grand_total), inv.payment_status.value,
        ])

    auto_col_width(ws)
    date_range = f"{from_date or 'all'}_{to_date or 'all'}"
    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"sales_register_{date_range}.xlsx")


# ── Cash Book Excel ───────────────────────────────────────────

@router.get("/cashbook-excel")
async def export_cashbook_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Export cash book to Excel."""
    tenant_id = payload["tenant_id"]
    stmt = (
        select(CashEntry)
        .where(CashEntry.tenant_id == tenant_id)
        .order_by(CashEntry.entry_date, CashEntry.id)
    )
    if from_date:
        stmt = stmt.where(CashEntry.entry_date >= from_date)
    if to_date:
        stmt = stmt.where(CashEntry.entry_date <= to_date)

    result  = await db.execute(stmt)
    entries = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Book"

    headers = ["Date", "Type", "Description", "Cash In", "Cash Out", "Bank In", "Balance", "Reference"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    running = Decimal("0")
    for e in entries:
        amt   = Decimal(str(e.amount))
        etype = e.entry_type.value
        if etype in ("cash_in", "bank_in"):
            running += amt
        elif etype in ("cash_out", "cash_to_bank"):
            running -= amt

        ws.append([
            e.entry_date.isoformat(), etype, e.description or "",
            float(amt) if etype in ("cash_in", "bank_in") else 0,
            float(amt) if etype in ("cash_out", "cash_to_bank") else 0,
            float(amt) if etype == "bank_in" else 0,
            float(running),
            e.bank_reference or "",
        ])

    auto_col_width(ws)
    date_range = f"{from_date or 'all'}_{to_date or 'all'}"
    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"cash_book_{date_range}.xlsx")


# ── Item-wise Excel ───────────────────────────────────────────

@router.get("/itemwise-excel")
async def export_itemwise_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Export item-wise sales report to Excel."""
    tenant_id = payload["tenant_id"]
    stmt = (
        select(Invoice)
        .where(Invoice.tenant_id == tenant_id, Invoice.status == "active")
        .order_by(Invoice.invoice_date.desc())
    )
    if from_date:
        stmt = stmt.where(Invoice.invoice_date >= from_date)
    if to_date:
        stmt = stmt.where(Invoice.invoice_date <= to_date)

    result   = await db.execute(stmt)
    invoices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item-wise"

    headers = [
        "Invoice No", "Date", "Customer", "Mobile", "PAN", "Mode",
        "Category", "Purity", "Description", "Qty", "Unit", "Rate", "Making", "Amount"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for inv in invoices:
        items_result = await db.execute(
            select(InvoiceItem).where(InvoiceItem.invoice_id == inv.id)
        )
        for item in items_result.scalars():
            ws.append([
                inv.invoice_no, inv.invoice_date.isoformat(),
                inv.customer_name, inv.customer_mobile, inv.customer_pan or "",
                inv.pay_mode.value, item.category.value, item.purity or "",
                item.description, float(item.qty), item.unit.value,
                float(item.rate), float(item.making_charges), float(item.amount),
            ])

    auto_col_width(ws)
    date_range = f"{from_date or 'all'}_{to_date or 'all'}"
    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"itemwise_summary_{date_range}.xlsx")


# ── SFT Excel ─────────────────────────────────────────────────

@router.get("/sft-excel")
async def export_sft_excel(
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    """Export SFT register to Excel."""
    tenant_id = payload["tenant_id"]
    result    = await db.execute(
        select(Customer).where(Customer.tenant_id == tenant_id, Customer.sft_flagged == True)
    )
    customers = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFT"

    headers = ["Customer", "Mobile", "PAN", "Cash Receipts FY", "SFT Threshold", "PAN Missing"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for c in customers:
        ws.append([
            c.name, c.mobile, c.pan or "",
            float(c.cash_receipts_fy), float(SFT_THRESHOLD),
            "YES" if not c.pan else "No",
        ])

    auto_col_width(ws)
    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"sft_register_{date.today().isoformat()}.xlsx")


# ── Section 269ST Excel ───────────────────────────────────────

@router.get("/section-269st-excel")
async def export_section_269st_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Export Section 269ST violation register to Excel.
    Uses INVOICE-level data: cash invoices with grand_total >= Rs. 2,00,000.
    Section 269ST prohibits receiving Rs. 2L+ in cash in a single transaction.
    """
    from decimal import Decimal
    tenant_id = payload["tenant_id"]
    threshold = Decimal("200000")

    stmt = (
        select(Invoice)
        .where(
            Invoice.tenant_id   == tenant_id,
            Invoice.status      == "active",
            Invoice.pay_mode    == "Cash",
            Invoice.grand_total >= threshold,
        )
        .order_by(Invoice.invoice_date.desc())
    )
    if from_date:
        stmt = stmt.where(Invoice.invoice_date >= from_date)
    if to_date:
        stmt = stmt.where(Invoice.invoice_date <= to_date)

    result   = await db.execute(stmt)
    invoices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sec 269ST Violations"

    headers = [
        "Invoice Date", "Invoice No", "Customer Name", "Mobile", "PAN",
        "Cash Amount (Rs.)", "Penalty Risk (Rs.)", "Notes"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for inv in invoices:
        cust = await db.get(Customer, (inv.customer_mobile, tenant_id))
        pan  = inv.customer_pan or (cust.pan if cust else "MISSING")
        ws.append([
            inv.invoice_date.isoformat(), inv.invoice_no,
            inv.customer_name, inv.customer_mobile, pan,
            float(inv.grand_total), float(inv.grand_total),
            inv.notes or "",
        ])

    auto_col_width(ws)
    date_range = f"{from_date or 'all'}_{to_date or 'all'}"

    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"section_269st_violations_{date_range}.xlsx")

# ── GSTR-1 Excel ──────────────────────────────────────────────
# Issue 4 fix — endpoint was missing, frontend Excel button returned 404

@router.get("/gstr1-excel")
async def export_gstr1_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Export GSTR-1 register to Excel."""
    tenant_id = payload["tenant_id"]
    stmt = (
        select(Invoice)
        .where(Invoice.tenant_id == tenant_id, Invoice.status == "active")
        .order_by(Invoice.invoice_date.desc())
    )
    if from_date:
        stmt = stmt.where(Invoice.invoice_date >= from_date)
    if to_date:
        stmt = stmt.where(Invoice.invoice_date <= to_date)

    result   = await db.execute(stmt)
    invoices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GSTR-1"

    headers = [
        "Invoice No", "Date", "Customer Name", "GSTIN", "State", "HSN Code",
        "GST Type", "Taxable Value", "CGST Rate%", "CGST Amt", "SGST Rate%", "SGST Amt",
        "IGST Amt", "Grand Total"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for inv in invoices:
        ws.append([
            inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name,
            inv.customer_gstin or "Unregistered", inv.customer_state or "",
            "7113", inv.gst_type.value,
            float(inv.subtotal),
            float(inv.gst_rate / 2), float(inv.cgst),
            float(inv.gst_rate / 2), float(inv.sgst),
            float(inv.igst), float(inv.grand_total),
        ])

    auto_col_width(ws)
    date_range = f"{from_date or 'all'}_{to_date or 'all'}"
    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"gstr1_{date_range}.xlsx")


# ── Outstanding Register Excel ─────────────────────────────────
# Issue 4 fix — endpoint was missing

@router.get("/outstanding-excel")
async def export_outstanding_excel(
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    """Export outstanding balances register to Excel."""
    tenant_id = payload["tenant_id"]
    result = await db.execute(
        select(Invoice).where(
            Invoice.tenant_id      == tenant_id,
            Invoice.payment_status != "paid",
            Invoice.status         == "active",
        ).order_by(Invoice.invoice_date.desc())
    )
    invoices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding"

    headers = ["Invoice No", "Date", "Customer Name", "Mobile", "Grand Total", "Amount Paid", "Outstanding"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for inv in invoices:
        ws.append([
            inv.invoice_no, inv.invoice_date.isoformat(),
            inv.customer_name, inv.customer_mobile,
            float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding),
        ])

    auto_col_width(ws)
    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"outstanding_register_{date.today().isoformat()}.xlsx")


# ── FIFO Valuation Excel ───────────────────────────────────────
# Issue 4 fix — endpoint was missing

@router.get("/fifo-excel")
async def export_fifo_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Export FIFO stock valuation to Excel."""
    from utils.business import fifo_valuation
    from models import StockTransaction

    tenant_id = payload["tenant_id"]
    cutoff    = date.today()

    stocks_result = await db.execute(
        select(StockItem).where(
            StockItem.tenant_id == tenant_id,
            StockItem.category  != "Polish Charges",
            StockItem.is_active == True,
        )
    )
    stocks = stocks_result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIFO Valuation"

    headers = ["Category", "Purity", "Description", "Unit", "Qty on Hand", "Avg Rate", "Total Value"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    grand_total = 0.0
    for stock in stocks:
        txns_result = await db.execute(
            select(StockTransaction).where(
                StockTransaction.stock_item_id == stock.id,
                StockTransaction.txn_date     <= cutoff,
                StockTransaction.txn_type.in_(["purchase", "opening"]),
                StockTransaction.lot_remaining > 0,
            ).order_by(StockTransaction.txn_date)
        )
        lots = [
            {"qty_remaining": t.lot_remaining, "purchase_rate": t.purchase_rate}
            for t in txns_result.scalars()
            if t.purchase_rate is not None
        ]
        valuation = fifo_valuation(lots)
        grand_total += float(valuation["total_value"])

        ws.append([
            stock.category.value, stock.purity or "", stock.description,
            stock.unit.value, float(stock.qty_on_hand),
            float(valuation["avg_rate"]), float(valuation["total_value"]),
        ])

    # Total row
    ws.append(["", "", "", "TOTAL", "", "", grand_total])
    auto_col_width(ws)
    # Add Account Register and Dashboard sheets to every Excel export
    await add_account_sheet(wb, db, tenant_id, from_date, to_date)
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    # Move Dashboard to first position, Account Register to second
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Account Register", offset=-len(wb.sheetnames)+2)
    return _stream_workbook(wb, f"fifo_valuation_{cutoff.isoformat()}.xlsx")


# ── All Reports Excel (combined workbook) ──────────────────────
# Issue 4 fix — "All Reports Excel" button called this endpoint which was missing


# ── Supplier Excel Export ────────────────────────────────────

@router.get("/supplier-invoices-excel")
async def export_supplier_invoices_excel(
    mobile:    Optional[str]  = Query(None),
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Supplier purchase invoices Excel export."""
    tid  = payload["tenant_id"]
    stmt = select(SupplierInvoice).where(SupplierInvoice.tenant_id == tid, SupplierInvoice.status == "active")
    if mobile:    stmt = stmt.where(SupplierInvoice.supplier_mobile == mobile)
    if from_date: stmt = stmt.where(SupplierInvoice.invoice_date >= from_date)
    if to_date:   stmt = stmt.where(SupplierInvoice.invoice_date <= to_date)
    r    = await db.execute(stmt.order_by(SupplierInvoice.invoice_date.desc()))
    invs = r.scalars().all()

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    add_sheet(wb, "Supplier Invoices", [
        "Invoice No", "Date", "Supplier Name", "Mobile",
        "Subtotal", "CGST", "SGST", "IGST", "Grand Total",
        "Amount Paid", "Outstanding", "Status",
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.supplier_name, inv.supplier_mobile,
         float(inv.subtotal), float(inv.cgst), float(inv.sgst), float(inv.igst),
         float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding),
         inv.payment_status.value if hasattr(inv.payment_status,"value") else str(inv.payment_status)]
        for inv in invs
    ])
    fname = f"supplier_invoices_{from_date or 'all'}.xlsx"
    return _stream_workbook(wb, fname)


@router.get("/supplier-payments-excel")
async def export_supplier_payments_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    tid  = payload["tenant_id"]
    stmt = select(SupplierPayment).where(SupplierPayment.tenant_id == tid)
    if from_date: stmt = stmt.where(SupplierPayment.payment_date >= from_date)
    if to_date:   stmt = stmt.where(SupplierPayment.payment_date <= to_date)
    r    = await db.execute(stmt.order_by(SupplierPayment.payment_date.desc()))
    pays = r.scalars().all()

    rows = []
    for p in pays:
        sup = await db.get(Supplier, (p.supplier_mobile, tid))
        rows.append([
            p.payment_date.isoformat(), sup.name if sup else "—", p.supplier_mobile,
            float(p.amount), p.pay_mode, p.reference_no or "—", p.notes or "",
        ])

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    add_sheet(wb, "Supplier Payments", [
        "Date", "Supplier Name", "Mobile", "Amount", "Mode", "Reference", "Notes"
    ], rows)
    return _stream_workbook(wb, f"supplier_payments_{from_date or 'all'}.xlsx")


@router.get("/supplier-advances-excel")
async def export_supplier_advances_excel(
    payload: dict        = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    """Download all supplier advances as Excel."""
    from models import SupplierAdvance, Supplier
    tid = payload["tenant_id"]
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Supplier Advances"

    headers = ["Date", "Supplier Name", "Mobile", "Amount (₹)", "Remaining (₹)", "Mode", "Notes"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    result = await db.execute(
        select(SupplierAdvance)
        .where(SupplierAdvance.tenant_id == tid)
        .order_by(SupplierAdvance.advance_date.desc())
    )
    for a in result.scalars().all():
        sup = await db.get(Supplier, (a.supplier_mobile, tid))
        ws.append([
            a.advance_date.isoformat(),
            sup.name if sup else "—",
            a.supplier_mobile,
            float(a.amount),
            float(a.remaining),
            a.pay_mode,
            a.notes or "",
        ])

    auto_col_width(ws)
    return _stream_workbook(wb, f"supplier_advances.xlsx")



@router.get("/supplier-ledger-excel")
async def export_supplier_ledger_excel(
    mobile:  str           = Query(...),
    payload: dict          = Depends(get_current_user_payload),
    db:      AsyncSession  = Depends(get_db),
):
    """Download a single supplier's complete ledger as a formatted Excel file."""
    from decimal import Decimal

    tid = payload["tenant_id"]
    today_dt = date.today()

    # Fetch supplier details
    sup = await db.get(Supplier, (mobile, tid))
    if not sup:
        from fastapi import HTTPException
        raise HTTPException(404, "Supplier not found")

    # Fetch invoices
    inv_r = await db.execute(
        select(SupplierInvoice)
        .where(SupplierInvoice.tenant_id == tid,
               SupplierInvoice.supplier_mobile == mobile,
               SupplierInvoice.status == "active")
        .order_by(SupplierInvoice.invoice_date)
    )
    invoices = inv_r.scalars().all()

    # Fetch payments
    pay_r = await db.execute(
        select(SupplierPayment)
        .where(SupplierPayment.tenant_id == tid,
               SupplierPayment.supplier_mobile == mobile)
        .order_by(SupplierPayment.payment_date)
    )
    payments = pay_r.scalars().all()

    # Fetch advances
    adv_r = await db.execute(
        select(SupplierAdvance)
        .where(SupplierAdvance.tenant_id == tid,
               SupplierAdvance.supplier_mobile == mobile)
        .order_by(SupplierAdvance.advance_date)
    )
    advances = adv_r.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Ledger"

    # ── Styles ────────────────────────────────────────────────
    dark_fill  = PatternFill("solid", fgColor="1A1A2E")
    gold_fill  = PatternFill("solid", fgColor="B8860B")
    hdr_fill   = PatternFill("solid", fgColor="2B2B4B")
    grn_fill   = PatternFill("solid", fgColor="1A3A1A")
    red_fill   = PatternFill("solid", fgColor="3A1A1A")
    white_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    gold_font  = Font(name="Calibri", bold=True, color="FFD700", size=14)
    lbl_font   = Font(name="Calibri", bold=True, color="AAAAAA", size=9)
    val_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    grn_font   = Font(name="Calibri", bold=True, color="66FF88", size=11)
    red_font   = Font(name="Calibri", bold=True, color="FF6666", size=11)
    ctr        = Alignment(horizontal="center", vertical="center")
    lft        = Alignment(horizontal="left",   vertical="center")
    rgt        = Alignment(horizontal="right",  vertical="center")

    def _c(row, col, val, font=None, fill=None, align=None, fmt=None):
        cell = ws.cell(row=row, column=col, value=val)
        if font:  cell.font   = font
        if fill:  cell.fill   = fill
        if align: cell.alignment = align
        if fmt:   cell.number_format = fmt
        return cell

    # ── Header block ──────────────────────────────────────────
    ws.merge_cells("A1:F1")
    _c(1, 1, f"SUPPLIER LEDGER — {sup.name.upper()}", font=gold_font, fill=dark_fill, align=ctr)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    _c(2, 1, f"Mobile: {mobile}  |  GSTIN: {sup.gstin or '—'}  |  PAN: {sup.pan or '—'}  |  Generated: {today_dt.strftime('%d %b %Y')}",
       font=Font(name="Calibri", color="AAAAAA", size=9),  fill=dark_fill, align=ctr)
    ws.row_dimensions[2].height = 18

    # ── Summary KPIs ──────────────────────────────────────────
    total_invoiced   = sum(float(i.grand_total)  for i in invoices)
    total_paid_inv   = sum(float(i.amount_paid)  for i in invoices)
    total_outstanding= sum(float(i.outstanding)  for i in invoices)
    total_payments   = sum(float(p.amount)       for p in payments)
    total_advances   = sum(float(a.amount)       for a in advances)

    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 22

    kpi_labels = ["Total Invoiced", "Amount Paid", "Outstanding", "Direct Payments", "Advances Given"]
    kpi_values = [total_invoiced, total_paid_inv, total_outstanding, total_payments, total_advances]
    for col, (lbl, val) in enumerate(zip(kpi_labels, kpi_values), 1):
        _c(4, col, lbl,  font=lbl_font, fill=hdr_fill, align=ctr)
        _c(5, col, val,  font=val_font if col != 3 else Font(name="Calibri",bold=True,color="FFAA44",size=11),
           fill=dark_fill, align=ctr, fmt="#,##0.00")

    ws.row_dimensions[6].height = 8

    # ── Ledger entries ────────────────────────────────────────
    # Build combined chronological ledger
    entries = []
    for inv in invoices:
        entries.append({
            "date": inv.invoice_date, "type": "Purchase Invoice",
            "ref":  inv.invoice_no or f"INV-{inv.id}",
            "debit": float(inv.grand_total), "credit": 0.0,
            "notes": inv.notes or ""
        })
    for p in payments:
        entries.append({
            "date": p.payment_date, "type": "Payment",
            "ref":  p.reference_no or f"PAY-{p.id}",
            "debit": 0.0, "credit": float(p.amount),
            "notes": p.notes or ""
        })
    for a in advances:
        entries.append({
            "date": a.advance_date, "type": "Advance",
            "ref":  f"ADV-{a.id}",
            "debit": 0.0, "credit": float(a.amount),
            "notes": a.notes or ""
        })
    entries.sort(key=lambda x: x["date"])

    # Table header
    ws.row_dimensions[7].height = 20
    col_headers = ["Date", "Type", "Reference", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
    for col, hdr in enumerate(col_headers, 1):
        _c(7, col, hdr, font=white_font, fill=gold_fill, align=ctr)

    # Table rows
    running = Decimal("0")
    for i, e in enumerate(entries):
        row = 8 + i
        ws.row_dimensions[row].height = 17
        debit  = Decimal(str(e["debit"]))
        credit = Decimal(str(e["credit"]))
        running = running + debit - credit
        is_inv = e["type"] == "Purchase Invoice"
        row_fill = PatternFill("solid", fgColor="1A1A2E" if i % 2 == 0 else "151528")
        _c(row, 1, e["date"].isoformat(), font=Font(name="Calibri",color="CCCCCC",size=10), fill=row_fill, align=ctr)
        _c(row, 2, e["type"],             font=Font(name="Calibri",bold=True,
           color="FF9944" if is_inv else "66CC88",size=10), fill=row_fill, align=lft)
        _c(row, 3, e["ref"],              font=Font(name="Calibri",color="DDDDDD",size=10), fill=row_fill, align=lft)
        _c(row, 4, float(debit)  if debit  > 0 else "",
           font=Font(name="Calibri",bold=True,color="FF6666",size=10), fill=row_fill, align=rgt,
           fmt="#,##0.00" if debit > 0 else None)
        _c(row, 5, float(credit) if credit > 0 else "",
           font=Font(name="Calibri",bold=True,color="66FF88",size=10), fill=row_fill, align=rgt,
           fmt="#,##0.00" if credit > 0 else None)
        bal_color = "FFAA44" if float(running) > 0 else "66FF88"
        _c(row, 6, float(running),
           font=Font(name="Calibri",bold=True,color=bal_color,size=10), fill=row_fill, align=rgt, fmt="#,##0.00")

    # ── Footer ────────────────────────────────────────────────
    last_row = 8 + len(entries)
    ws.row_dimensions[last_row].height = 8
    ws.merge_cells(f"A{last_row+1}:F{last_row+1}")
    bal_label = "AMOUNT DUE TO SUPPLIER" if float(running) > 0 else "CREDIT BALANCE"
    _c(last_row+1, 1, f"{bal_label}: ₹{float(running):,.2f}",
       font=Font(name="Calibri",bold=True,color="FFD700",size=12),
       fill=PatternFill("solid",fgColor="0D0D1A"), align=ctr)
    ws.row_dimensions[last_row+1].height = 24

    # Column widths
    for col, w in enumerate([14, 20, 20, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    safe_name = sup.name.replace(" ", "_")[:20]
    filename  = f"Ledger_{safe_name}_{today_dt.isoformat()}.xlsx"
    return _stream_workbook(wb, filename)


# ── Customer Ledger Excel ─────────────────────────────────

@router.get("/customer-ledger-excel")
async def export_customer_ledger_excel(
    mobile:  str           = Query(...),
    payload: dict          = Depends(get_current_user_payload),
    db:      AsyncSession  = Depends(get_db),
):
    """Download a single customer's complete ledger as a formatted Excel file."""
    from fastapi import HTTPException

    tid      = payload["tenant_id"]
    today_dt = date.today()

    cust_r = await db.execute(
        select(Customer).where(Customer.tenant_id == tid, Customer.mobile == mobile)
    )
    cust = cust_r.scalars().first()
    if not cust:
        raise HTTPException(404, "Customer not found")

    inv_r = await db.execute(
        select(Invoice)
        .where(Invoice.tenant_id == tid, Invoice.customer_mobile == mobile,
               Invoice.status == "active")
        .order_by(Invoice.invoice_date)
    )
    invoices = inv_r.scalars().all()

    pay_r = await db.execute(
        select(Payment)
        .where(Payment.tenant_id == tid, Payment.customer_mobile == mobile)
        .order_by(Payment.payment_date)
    )
    payments = pay_r.scalars().all()

    adv_r = await db.execute(
        select(Advance)
        .where(Advance.tenant_id == tid, Advance.customer_mobile == mobile)
        .order_by(Advance.advance_date)
    )
    advances = adv_r.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"

    # Light theme — matches the new Excel colour scheme
    navy_fill   = PatternFill("solid", fgColor="1F3864")
    gold_fill   = PatternFill("solid", fgColor="C8900A")
    hdr_fill    = PatternFill("solid", fgColor="E8F0FE")
    alt_fill    = PatternFill("solid", fgColor="F7F9FF")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    red_fill    = PatternFill("solid", fgColor="FFEBEE")
    green_fill  = PatternFill("solid", fgColor="E8F5E9")
    title_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
    sub_font    = Font(name="Calibri", color="AAAAAA", size=9)
    lbl_font    = Font(name="Calibri", bold=True, color="555555", size=9)
    val_font    = Font(name="Calibri", bold=True, color="1A237E", size=11)
    grn_font    = Font(name="Calibri", bold=True, color="1B5E20", size=11)
    red_font    = Font(name="Calibri", bold=True, color="B71C1C", size=11)
    body_font   = Font(name="Calibri", size=10)
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center", indent=1)
    rgt  = Alignment(horizontal="right",  vertical="center")
    thin = Border(left=Side(style="thin",color="E0E0E0"),right=Side(style="thin",color="E0E0E0"),
                  top=Side(style="thin",color="E0E0E0"),bottom=Side(style="thin",color="E0E0E0"))

    def _c(row, col, val, font=None, fill=None, align=None, fmt=None):
        cell = ws.cell(row=row, column=col, value=val)
        if font:  cell.font   = font
        if fill:  cell.fill   = fill
        if align: cell.alignment = align
        if fmt:   cell.number_format = fmt
        cell.border = thin
        return cell

    # Title
    ws.merge_cells("A1:F1")
    _c(1,1, f"CUSTOMER LEDGER — {cust.name.upper()}",
       font=title_font, fill=navy_fill, align=ctr)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:F2")
    _c(2,1, f"Mobile: {mobile}  |  PAN: {cust.pan or '—'}  |  Generated: {today_dt.strftime('%d %b %Y')}",
       font=sub_font, fill=hdr_fill, align=ctr)
    ws.row_dimensions[2].height = 18

    # KPI row
    total_invoiced    = sum(float(i.grand_total)  for i in invoices)
    total_paid_inv    = sum(float(i.amount_paid)  for i in invoices)
    total_outstanding = sum(float(i.outstanding)  for i in invoices)
    total_payments    = sum(float(p.amount)        for p in payments)
    total_advances    = sum(float(a.amount)        for a in advances)

    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 22

    kpi_labels = ["Total Invoiced","Amount Received","Outstanding","Direct Payments","Advances Received"]
    kpi_values = [total_invoiced, total_paid_inv, total_outstanding, total_payments, total_advances]
    for col, (lbl, val) in enumerate(zip(kpi_labels, kpi_values), 1):
        _c(4, col, lbl, font=lbl_font, fill=hdr_fill, align=ctr)
        vf = red_font if col == 3 and val > 0 else (grn_font if col == 3 else val_font)
        vfill = red_fill if col == 3 and val > 0 else (green_fill if col == 3 else white_fill)
        _c(5, col, val, font=vf, fill=vfill, align=ctr, fmt="#,##0.00")

    ws.row_dimensions[6].height = 6

    # Build ledger entries
    entries = []
    for inv in invoices:
        entries.append({"date": inv.invoice_date, "type": "Invoice",
                         "ref": inv.invoice_no or f"INV-{inv.id}",
                         "debit": float(inv.grand_total), "credit": 0.0,
                         "notes": inv.notes or ""})
    for p in payments:
        entries.append({"date": p.payment_date, "type": "Payment",
                         "ref": p.reference_no or f"PAY-{p.id}",
                         "debit": 0.0, "credit": float(p.amount),
                         "notes": p.notes or ""})
    for a in advances:
        entries.append({"date": a.advance_date, "type": "Advance",
                         "ref": f"ADV-{a.id}",
                         "debit": 0.0, "credit": float(a.amount),
                         "notes": a.notes or ""})
    entries.sort(key=lambda x: x["date"])

    # Header row
    ws.row_dimensions[7].height = 20
    col_hdrs = ["Date","Type","Reference","Debit (₹)","Credit (₹)","Balance (₹)"]
    for col, hdr in enumerate(col_hdrs, 1):
        _c(7, col, hdr, font=Font(name="Calibri",bold=True,color="FFFFFF",size=10),
           fill=gold_fill, align=ctr)

    # Ledger rows
    running = Decimal("0")
    for i, e in enumerate(entries):
        row  = 8 + i
        ws.row_dimensions[row].height = 17
        debit  = Decimal(str(e["debit"]))
        credit = Decimal(str(e["credit"]))
        running = running + debit - credit
        rf = alt_fill if i % 2 == 0 else white_fill
        is_inv = e["type"] == "Invoice"
        _c(row,1,e["date"].isoformat(),
           font=Font(name="Calibri",color="555555",size=10), fill=rf, align=ctr)
        _c(row,2,e["type"],
           font=Font(name="Calibri",bold=True,
                     color="1565C0" if is_inv else "2E7D32",size=10), fill=rf, align=lft)
        _c(row,3,e["ref"],
           font=Font(name="Calibri",color="333333",size=10), fill=rf, align=lft)
        _c(row,4,float(debit)  if debit  > 0 else "",
           font=Font(name="Calibri",bold=True,color="C62828",size=10), fill=rf, align=rgt,
           fmt="#,##0.00" if debit > 0 else None)
        _c(row,5,float(credit) if credit > 0 else "",
           font=Font(name="Calibri",bold=True,color="2E7D32",size=10), fill=rf, align=rgt,
           fmt="#,##0.00" if credit > 0 else None)
        bal_color = "C62828" if float(running) > 0 else "1B5E20"
        _c(row,6,float(running),
           font=Font(name="Calibri",bold=True,color=bal_color,size=10),
           fill=red_fill if float(running) > 0 else green_fill, align=rgt, fmt="#,##0.00")

    # Footer
    last = 8 + len(entries)
    ws.row_dimensions[last].height = 8
    ws.merge_cells(f"A{last+1}:F{last+1}")
    bal_label = "AMOUNT RECEIVABLE" if float(running) > 0 else "CREDIT BALANCE"
    _c(last+1,1, f"{bal_label}: ₹{float(running):,.2f}",
       font=Font(name="Calibri",bold=True,color="FFFFFF",size=12),
       fill=navy_fill if float(running)>0 else PatternFill("solid",fgColor="2E7D32"), align=ctr)
    ws.row_dimensions[last+1].height = 24

    # Column widths + no gridlines
    for col, w in enumerate([14,20,22,16,16,16],1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.sheet_view.showGridLines = False

    safe = cust.name.replace(" ","_")[:20]
    return _stream_workbook(wb, f"Ledger_{safe}_{today_dt.isoformat()}.xlsx")


@router.get("/supplier-all-excel")
async def export_supplier_all_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """All-in-one Supplier Excel with Dashboard summary as first sheet."""
    from models import (Supplier, SupplierInvoice, SupplierInvoiceItem,
                        SupplierPayment, SupplierAdvance)
    from decimal import Decimal
    from datetime import date as _date
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tid      = payload["tenant_id"]
    today_dt = _date.today()
    wb       = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Fetch all data up front ──────────────────────────────
    sup_r = await db.execute(
        select(Supplier).where(Supplier.tenant_id == tid).order_by(Supplier.name)
    )
    sup_list = sup_r.scalars().all()

    inv_stmt = (
        select(SupplierInvoice)
        .where(SupplierInvoice.tenant_id == tid, SupplierInvoice.status == "active")
        .order_by(SupplierInvoice.invoice_date.desc())
    )
    if from_date: inv_stmt = inv_stmt.where(SupplierInvoice.invoice_date >= from_date)
    if to_date:   inv_stmt = inv_stmt.where(SupplierInvoice.invoice_date <= to_date)
    inv_r    = await db.execute(inv_stmt)
    sup_invs = inv_r.scalars().all()

    pay_r = await db.execute(
        select(SupplierPayment)
        .where(SupplierPayment.tenant_id == tid)
        .order_by(SupplierPayment.payment_date.desc())
    )
    sup_pays = pay_r.scalars().all()

    adv_r = await db.execute(
        select(SupplierAdvance)
        .where(SupplierAdvance.tenant_id == tid)
        .order_by(SupplierAdvance.advance_date.desc())
    )
    sup_advs = adv_r.scalars().all()

    # ── Compute summary figures ──────────────────────────────
    total_invoiced   = sum(float(inv.grand_total)   for inv in sup_invs)
    total_paid_inv   = sum(float(inv.amount_paid)   for inv in sup_invs)
    total_outstanding= sum(float(inv.outstanding)   for inv in sup_invs)
    total_payments   = sum(float(p.amount)          for p in sup_pays)
    total_advances   = sum(float(a.amount)          for a in sup_advs)
    adv_remaining    = sum(float(a.remaining if hasattr(a,"remaining") else a.amount) for a in sup_advs)
    inv_count_active = len([i for i in sup_invs if float(i.outstanding) > 0.005])

    # Top 5 suppliers by outstanding
    sup_out_map: dict = {}
    for inv in sup_invs:
        key = inv.supplier_mobile
        sup_out_map[key] = sup_out_map.get(key, {"name": inv.supplier_name or inv.supplier_mobile, "out": 0.0})
        sup_out_map[key]["out"] += float(inv.outstanding)
    top5 = sorted(sup_out_map.values(), key=lambda x: x["out"], reverse=True)[:5]

    # ── Sheet 1: Dashboard ───────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard")

    # Styles
    gold_fill  = PatternFill("solid", fgColor="B8860B")
    dark_fill  = PatternFill("solid", fgColor="1A1A2E")
    amber_fill = PatternFill("solid", fgColor="CC7722")
    green_fill = PatternFill("solid", fgColor="1E6B2E")
    red_fill   = PatternFill("solid", fgColor="8B0000")
    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    title_font = Font(name="Calibri", bold=True, color="FFD700", size=16)
    kpi_font   = Font(name="Calibri", bold=True, color="FFD700", size=14)
    sub_font   = Font(name="Calibri", color="CCCCCC", size=10)
    val_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border= Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )

    def dash_cell(row, col, value, font=None, fill=None, align="left", border=True):
        c = ws_dash.cell(row=row, column=col, value=value)
        if font:  c.font = font
        if fill:  c.fill = fill
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if border: c.border = thin_border
        return c

    # Title row
    ws_dash.merge_cells("A1:F1")
    dash_cell(1, 1, f"🏭  SUPPLIER DASHBOARD  —  {today_dt.strftime('%d %b %Y')}",
              font=title_font, fill=dark_fill, align="center", border=False)
    ws_dash.row_dimensions[1].height = 36

    # KPI headers row 3
    ws_dash.row_dimensions[3].height = 22
    kpi_headers = ["Total Suppliers", "Total Invoiced (₹)", "Total Paid (₹)",
                   "Total Outstanding (₹)", "Total Payments Made (₹)", "Advances Given (₹)"]
    kpi_values  = [len(sup_list), total_invoiced, total_paid_inv,
                   total_outstanding, total_payments, total_advances]
    for col, (hdr, val) in enumerate(zip(kpi_headers, kpi_values), 1):
        dash_cell(3, col, hdr, font=Font(name="Calibri", bold=True, color="FFFFFF", size=10),
                  fill=dark_fill, align="center")
        c = dash_cell(4, col,
                      len(sup_list) if col == 1 else f"₹{val:,.2f}",
                      font=kpi_font, fill=gold_fill, align="center")
        ws_dash.row_dimensions[4].height = 28

    # Blank row
    ws_dash.row_dimensions[5].height = 8

    # Section: Outstanding Invoices summary
    ws_dash.merge_cells("A6:F6")
    dash_cell(6, 1, "  OUTSTANDING INVOICES SUMMARY", font=hdr_font, fill=dark_fill, align="left")
    ws_dash.row_dimensions[6].height = 20

    dash_cell(7, 1, "Invoices with balance due", font=val_font, fill=dark_fill)
    dash_cell(7, 2, inv_count_active, font=kpi_font, fill=amber_fill, align="center")
    dash_cell(7, 3, "Total outstanding", font=val_font, fill=dark_fill)
    dash_cell(7, 4, f"₹{total_outstanding:,.2f}", font=kpi_font, fill=amber_fill, align="center")
    dash_cell(7, 5, "Advances remaining", font=val_font, fill=dark_fill)
    dash_cell(7, 6, f"₹{adv_remaining:,.2f}", font=kpi_font, fill=green_fill, align="center")
    ws_dash.row_dimensions[7].height = 24

    # Blank row
    ws_dash.row_dimensions[8].height = 8

    # Section: Top 5 suppliers by outstanding
    ws_dash.merge_cells("A9:F9")
    dash_cell(9, 1, "  TOP SUPPLIERS BY OUTSTANDING BALANCE", font=hdr_font, fill=dark_fill, align="left")
    ws_dash.row_dimensions[9].height = 20

    t5_headers = ["Rank", "Supplier Name", "Outstanding (₹)", "", "", ""]
    for col, h in enumerate(t5_headers, 1):
        dash_cell(10, col, h, font=Font(name="Calibri", bold=True, color="AAAAAA", size=9),
                  fill=dark_fill, align="center")
    ws_dash.row_dimensions[10].height = 18

    for rank, sup in enumerate(top5, 1):
        fill = red_fill if sup["out"] > 100000 else amber_fill
        dash_cell(10 + rank, 1, rank,             font=val_font, fill=dark_fill, align="center")
        dash_cell(10 + rank, 2, sup["name"],      font=val_font, fill=dark_fill)
        dash_cell(10 + rank, 3, f"₹{sup['out']:,.2f}", font=Font(name="Calibri", bold=True, color="FFD700", size=11), fill=fill, align="right")
        for col in range(4, 7):
            dash_cell(10 + rank, col, "", fill=dark_fill)
        ws_dash.row_dimensions[10 + rank].height = 22

    # Blank + sheet guide
    guide_row = 17
    ws_dash.merge_cells(f"A{guide_row}:F{guide_row}")
    dash_cell(guide_row, 1, "  SHEETS IN THIS WORKBOOK", font=hdr_font, fill=dark_fill, align="left")
    ws_dash.row_dimensions[guide_row].height = 20
    sheet_guide = [
        ("Dashboard",        "This summary page"),
        ("Suppliers",        "All supplier master records"),
        ("Purchase Invoices","All purchase invoices with GST breakdown"),
        ("Payments",         "All supplier payments recorded"),
        ("Advances",         "Advance payments to suppliers"),
        ("Outstanding",      "Invoices with pending balances"),
        ("GSTR-2B",          "Input GST reconciliation data"),
    ]
    for i, (sheet, desc) in enumerate(sheet_guide):
        dash_cell(guide_row + 1 + i, 1, sheet, font=Font(name="Calibri", bold=True, color="FFD700", size=10), fill=dark_fill)
        ws_dash.merge_cells(f"B{guide_row+1+i}:F{guide_row+1+i}")
        dash_cell(guide_row + 1 + i, 2, desc, font=sub_font, fill=dark_fill)
        ws_dash.row_dimensions[guide_row + 1 + i].height = 18

    # Column widths for dashboard
    for col, width in enumerate([20, 28, 22, 22, 26, 22], 1):
        ws_dash.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 2: Suppliers ───────────────────────────────────
    add_sheet(wb, "Suppliers",
        ["Name", "Mobile", "GSTIN", "PAN", "State", "Email", "Address"],
        [[s.name, s.mobile, s.gstin or "", s.pan or "", s.state or "",
          s.email or "", s.address or ""] for s in sup_list])

    # ── Sheet 3: Purchase Invoices ───────────────────────────
    add_sheet(wb, "Purchase Invoices",
        ["Invoice No", "Date", "Supplier Name", "Mobile",
         "Subtotal", "CGST", "SGST", "IGST", "Grand Total",
         "Amount Paid", "Outstanding", "Payment Status"],
        [[inv.invoice_no, inv.invoice_date.isoformat(), inv.supplier_name or "", inv.supplier_mobile,
          float(inv.subtotal), float(inv.cgst), float(inv.sgst), float(inv.igst),
          float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding),
         inv.payment_status.value if hasattr(inv.payment_status,"value") else str(inv.payment_status)]
         for inv in sup_invs])

    # ── Sheet 4: Payments ────────────────────────────────────
    add_sheet(wb, "Payments",
        ["Date", "Supplier Name", "Mobile", "Amount", "Mode", "Reference", "Notes"],
        [[p.payment_date.isoformat(),
          next((s.name for s in sup_list if s.mobile == p.supplier_mobile), p.supplier_mobile),
          p.supplier_mobile, float(p.amount), p.pay_mode if isinstance(p.pay_mode, str) else (p.pay_mode.value if hasattr(p.pay_mode,"value") else str(p.pay_mode)),
          p.reference_no or "", p.notes or ""]
         for p in sup_pays])

    # ── Sheet 5: Advances ────────────────────────────────────
    add_sheet(wb, "Advances",
        ["Date", "Supplier Name", "Mobile", "Amount", "Remaining", "Mode", "Notes"],
        [[a.advance_date.isoformat(),
          next((s.name for s in sup_list if s.mobile == a.supplier_mobile), a.supplier_mobile),
          a.supplier_mobile, float(a.amount),
          float(a.remaining if hasattr(a,"remaining") else a.amount),
          a.pay_mode if isinstance(a.pay_mode, str) else (a.pay_mode.value if hasattr(a.pay_mode,"value") else str(a.pay_mode)),
          a.notes or ""]
         for a in sup_advs])

    # ── Sheet 6: Outstanding ─────────────────────────────────
    outstanding_invs = [inv for inv in sup_invs if float(inv.outstanding) > 0.005]
    add_sheet(wb, "Outstanding",
        ["Invoice No", "Date", "Supplier Name", "Mobile",
         "Grand Total", "Amount Paid", "Outstanding"],
        [[inv.invoice_no, inv.invoice_date.isoformat(),
          inv.supplier_name or "", inv.supplier_mobile,
          float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding)]
         for inv in outstanding_invs])

    # ── Sheet 7: GSTR-2B ─────────────────────────────────────
    ws_gstr = wb.create_sheet("GSTR-2B")
    gstr_headers = ["Invoice No", "Date", "Supplier Name", "Supplier GSTIN",
                    "Taxable Value", "CGST", "SGST", "IGST", "Total GST", "Grand Total"]
    ws_gstr.append(gstr_headers)
    for inv in sup_invs:
        sup_obj = next((s for s in sup_list if s.mobile == inv.supplier_mobile), None)
        taxable = float(inv.subtotal)
        cgst = float(inv.cgst); sgst = float(inv.sgst); igst = float(inv.igst)
        ws_gstr.append([
            inv.invoice_no, inv.invoice_date.isoformat(),
            inv.supplier_name or "", sup_obj.gstin if sup_obj else "",
            taxable, cgst, sgst, igst, cgst + sgst + igst,
            float(inv.grand_total),
        ])
    auto_col_width(ws_gstr)

    return _stream_workbook(wb, f"GoldTrader_Supplier_All_{today_dt.isoformat()}.xlsx")


@router.get("/all-reports-excel")
async def export_all_reports_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """
    Single Excel workbook containing all report sheets.
    Used by the 'All Reports Excel' button on the Reports page.
    """
    from utils.business import current_fy, fifo_valuation, SFT_THRESHOLD
    from models import (StockTransaction, Advance,
                        Supplier, SupplierInvoice, SupplierInvoiceItem,
                        SupplierPayment, SupplierAdvance)
    from decimal import Decimal

    tid      = payload["tenant_id"]
    fy_start, fy_end = current_fy()
    today_dt = date.today()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # helper: date-filtered invoice query
    async def get_invoices(active_only=True):
        stmt = select(Invoice).where(Invoice.tenant_id == tid)
        if active_only:
            stmt = stmt.where(Invoice.status == "active")
        if from_date:
            stmt = stmt.where(Invoice.invoice_date >= from_date)
        if to_date:
            stmt = stmt.where(Invoice.invoice_date <= to_date)
        r = await db.execute(stmt.order_by(Invoice.invoice_date.desc()))
        return r.scalars().all()

    invoices = await get_invoices()

    # ── Pre-fetch supplier data for dashboard ────────────────
    s_inv_r  = await db.execute(select(SupplierInvoice).where(
        SupplierInvoice.tenant_id == tid, SupplierInvoice.status == "active"))
    s_invs   = s_inv_r.scalars().all()
    s_pay_r  = await db.execute(select(SupplierPayment).where(SupplierPayment.tenant_id == tid))
    s_pays   = s_pay_r.scalars().all()
    s_adv_r  = await db.execute(select(SupplierAdvance).where(SupplierAdvance.tenant_id == tid))
    s_advs   = s_adv_r.scalars().all()
    sup_r    = await db.execute(select(Supplier).where(Supplier.tenant_id == tid))
    sup_list = sup_r.scalars().all()
    cust_r   = await db.execute(select(Customer).where(Customer.tenant_id == tid))
    cust_list= cust_r.scalars().all()

    # ── Customer summary figures ──────────────────────────────
    c_total_inv   = sum(float(i.grand_total)  for i in invoices)
    c_total_paid  = sum(float(i.amount_paid)  for i in invoices)
    c_outstanding = sum(float(i.outstanding)  for i in invoices)
    c_cash_sales  = sum(float(i.grand_total)  for i in invoices if hasattr(i.pay_mode,'value') and i.pay_mode.value=='Cash')
    c_inv_count   = len(invoices)

    # ── Supplier summary figures ──────────────────────────────
    s_total_inv   = sum(float(i.grand_total)  for i in s_invs)
    s_total_paid  = sum(float(i.amount_paid)  for i in s_invs)
    s_outstanding = sum(float(i.outstanding)  for i in s_invs)
    s_total_pays  = sum(float(p.amount)       for p in s_pays)
    s_total_advs  = sum(float(a.amount)       for a in s_advs)
    s_adv_remain  = sum(float(a.remaining)    for a in s_advs)

    # ── Sheet 0: Dashboard — light theme + compliance ──────────
    ws_d = wb.create_sheet("Dashboard", 0)

    # Light palette
    F_NAVY   = PatternFill("solid", fgColor="1F3864")   # navy — header bar only
    F_LBLUE  = PatternFill("solid", fgColor="DDEEFF")   # light blue — customer col
    F_LGREEN = PatternFill("solid", fgColor="E6F4EA")   # light green — supplier col
    F_WHITE  = PatternFill("solid", fgColor="FFFFFF")
    F_LGREY  = PatternFill("solid", fgColor="F7F7F7")
    F_LGOLD  = PatternFill("solid", fgColor="FFFDE7")   # very pale gold — totals
    F_RED_L  = PatternFill("solid", fgColor="FFEBEE")   # light red — outstanding/risk
    F_AMB_L  = PatternFill("solid", fgColor="FFF8E1")   # light amber — compliance
    F_HDR_C  = PatternFill("solid", fgColor="1565C0")   # customer section header
    F_HDR_S  = PatternFill("solid", fgColor="2E7D32")   # supplier section header
    F_HDR_K  = PatternFill("solid", fgColor="BF360C")   # compliance header

    T_TITLE  = Font(name="Calibri", bold=True,  color="FFFFFF", size=16)
    T_PERIOD = Font(name="Calibri", bold=False, color="555555", size=9,  italic=True)
    T_SHDR   = Font(name="Calibri", bold=True,  color="FFFFFF", size=11)
    T_LBL    = Font(name="Calibri", bold=False, color="444444", size=10)
    T_VAL    = Font(name="Calibri", bold=True,  color="1A237E", size=11)  # dark navy val
    T_VAL_G  = Font(name="Calibri", bold=True,  color="1B5E20", size=11)  # dark green
    T_VAL_R  = Font(name="Calibri", bold=True,  color="B71C1C", size=11)  # dark red
    T_COMP   = Font(name="Calibri", bold=True,  color="BF360C", size=10)  # compliance

    _ctr = Alignment(horizontal="center", vertical="center")
    _lft = Alignment(horizontal="left",   vertical="center", indent=1)
    _rgt = Alignment(horizontal="right",  vertical="center")
    _bdr = Border(left=Side(style="thin", color="CCCCCC"),
                  right=Side(style="thin", color="CCCCCC"),
                  top=Side(style="thin", color="CCCCCC"),
                  bottom=Side(style="thin", color="CCCCCC"))

    def _D(r, c, v, font=None, fill=None, align=None, fmt=None):
        cell = ws_d.cell(row=r, column=c, value=v)
        if font:  cell.font   = font
        if fill:  cell.fill   = fill
        if align: cell.alignment = align
        if fmt:   cell.number_format = fmt
        cell.border = _bdr
        return cell

    per        = today_dt.strftime("%d %b %Y")
    period_txt = ("Period: " +
                  (from_date.strftime("%d %b %Y") if from_date else "All time") +
                  "  to  " +
                  (to_date.strftime("%d %b %Y") if to_date else per))

    # ── R1: Title ────────────────────────────────────────────
    ws_d.merge_cells("A1:J1")
    _D(1,1,"GoldTrader Pro  —  Complete Business Dashboard",
       font=T_TITLE, fill=F_NAVY, align=_ctr)
    ws_d.row_dimensions[1].height = 34

    ws_d.merge_cells("A2:J2")
    _D(2,1,period_txt, font=T_PERIOD, fill=F_LGREY, align=_ctr)
    ws_d.row_dimensions[2].height = 16
    ws_d.row_dimensions[3].height = 4   # spacer

    # ── R4: Section headers ──────────────────────────────────
    ws_d.merge_cells("A4:E4")
    _D(4,1,"CUSTOMER (SALES) SUMMARY", font=T_SHDR, fill=F_HDR_C, align=_ctr)
    ws_d.merge_cells("F4:J4")
    _D(4,6,"SUPPLIER (PURCHASE) SUMMARY", font=T_SHDR, fill=F_HDR_S, align=_ctr)
    ws_d.row_dimensions[4].height = 22

    # ── R5–10: KPI rows ──────────────────────────────────────
    c_kpi = [
        ("Total Customers",      len(cust_list), False),
        ("Sales Invoices",       c_inv_count,    False),
        ("Total Sales (Rs)",     c_total_inv,    False),
        ("Total Collected (Rs)", c_total_paid,   False),
        ("Receivables (Rs)",     c_outstanding,  True),
        ("Cash Sales (Rs)",      c_cash_sales,   False),
    ]
    s_kpi = [
        ("Total Suppliers",       len(sup_list),  False),
        ("Purchase Invoices",     len(s_invs),    False),
        ("Total Purchases (Rs)",  s_total_inv,    False),
        ("Total Paid (Rs)",       s_total_pays,   False),
        ("Payables (Rs)",         s_outstanding,  True),
        ("Advances Given (Rs)",   s_total_advs,   False),
    ]
    for idx2,(lbl,val,flag) in enumerate(c_kpi):
        r = 5 + idx2
        ws_d.merge_cells(f"A{r}:B{r}")
        _D(r,1,lbl, font=T_LBL, fill=F_LBLUE, align=_lft)
        ws_d.merge_cells(f"C{r}:E{r}")
        vfill = F_RED_L if flag else F_WHITE
        vfont = T_VAL_R if flag else T_VAL
        fmt   = "#,##0" if isinstance(val,int) else "#,##0.00"
        _D(r,3,val, font=vfont, fill=vfill, align=_rgt, fmt=fmt)
        ws_d.row_dimensions[r].height = 20

    for idx2,(lbl,val,flag) in enumerate(s_kpi):
        r = 5 + idx2
        ws_d.merge_cells(f"F{r}:G{r}")
        _D(r,6,lbl, font=T_LBL, fill=F_LGREEN, align=_lft)
        ws_d.merge_cells(f"H{r}:J{r}")
        vfill = F_RED_L if flag else F_WHITE
        vfont = T_VAL_R if flag else T_VAL_G
        fmt   = "#,##0" if isinstance(val,int) else "#,##0.00"
        _D(r,8,val, font=vfont, fill=vfill, align=_rgt, fmt=fmt)
        ws_d.row_dimensions[r].height = 20

    ws_d.row_dimensions[11].height = 6   # spacer

    # ── R12: Top-5 section headers ───────────────────────────
    ws_d.merge_cells("A12:E12")
    _D(12,1,"TOP CUSTOMERS — BY RECEIVABLES",
       font=T_SHDR, fill=F_HDR_C, align=_ctr)
    ws_d.merge_cells("F12:J12")
    _D(12,6,"TOP SUPPLIERS — BY PAYABLES",
       font=T_SHDR, fill=F_HDR_S, align=_ctr)
    ws_d.row_dimensions[12].height = 20

    # Customer top-5
    cust_out_map: dict = {}
    for inv in invoices:
        k = inv.customer_mobile
        if k not in cust_out_map:
            cust_out_map[k] = {"name": inv.customer_name or k, "out": 0.0}
        cust_out_map[k]["out"] += float(inv.outstanding)
    top_c = sorted(cust_out_map.values(), key=lambda x: x["out"], reverse=True)[:5]
    for idx2, x in enumerate(top_c):
        r = 13 + idx2
        ws_d.merge_cells(f"A{r}:C{r}")
        _D(r,1,x["name"], font=T_LBL, fill=F_LGREY, align=_lft)
        ws_d.merge_cells(f"D{r}:E{r}")
        _D(r,4,x["out"],
           font=T_VAL_R if x["out"]>0 else T_VAL_G,
           fill=F_RED_L if x["out"]>0 else F_WHITE,
           align=_rgt, fmt="#,##0.00")
        ws_d.row_dimensions[r].height = 18

    # Supplier top-5
    sup_out_map2: dict = {}
    for inv in s_invs:
        k = inv.supplier_mobile
        if k not in sup_out_map2:
            sup_out_map2[k] = {"name": inv.supplier_name or k, "out": 0.0}
        sup_out_map2[k]["out"] += float(inv.outstanding)
    top_s = sorted(sup_out_map2.values(), key=lambda x: x["out"], reverse=True)[:5]
    for idx2, x in enumerate(top_s):
        r = 13 + idx2
        ws_d.merge_cells(f"F{r}:H{r}")
        _D(r,6,x["name"], font=T_LBL, fill=F_LGREY, align=_lft)
        ws_d.merge_cells(f"I{r}:J{r}")
        _D(r,9,x["out"],
           font=T_VAL_R if x["out"]>0 else T_VAL_G,
           fill=F_RED_L if x["out"]>0 else F_WHITE,
           align=_rgt, fmt="#,##0.00")
        ws_d.row_dimensions[r].height = 18

    ws_d.row_dimensions[18].height = 6   # spacer

    # ── R19+: Non-Compliance Summary ─────────────────────────
    ws_d.merge_cells("A19:J19")
    _D(19,1,"NON-COMPLIANCE SUMMARY",
       font=T_SHDR, fill=F_HDR_K, align=_ctr)
    ws_d.row_dimensions[19].height = 22

    # Sub-header
    for col, (hdr, merge_end) in enumerate([
        ("Compliance Area", "C"), ("Description", "G"),
        ("Count", "H"), ("Amount (Rs)", "J")
    ], start=1):
        pass  # we use merge below
    ws_d.merge_cells("A20:C20")
    _D(20,1,"Compliance Area",  font=Font(name="Calibri",bold=True,color="FFFFFF",size=10),
       fill=F_HDR_K, align=_lft)
    ws_d.merge_cells("D20:G20")
    _D(20,4,"Description",       font=Font(name="Calibri",bold=True,color="FFFFFF",size=10),
       fill=F_HDR_K, align=_lft)
    _D(20,8,"Count",             font=Font(name="Calibri",bold=True,color="FFFFFF",size=10),
       fill=F_HDR_K, align=_ctr)
    ws_d.merge_cells("I20:J20")
    _D(20,9,"Amount (Rs)",       font=Font(name="Calibri",bold=True,color="FFFFFF",size=10),
       fill=F_HDR_K, align=_rgt)
    ws_d.row_dimensions[20].height = 20

    # Compute compliance figures
    # 1) Sec 269ST — cash invoices >= 2,00,000
    s269_invs = [inv for inv in invoices
                 if hasattr(inv.pay_mode,"value")
                 and inv.pay_mode.value == "Cash"
                 and float(inv.grand_total) >= 200000]
    s269_cnt  = len(s269_invs)
    s269_amt  = sum(float(i.grand_total) for i in s269_invs)

    # 2) Cash-out > 10,000 from cash book
    ce_r2 = await db.execute(select(CashEntry).where(CashEntry.tenant_id == tid))
    all_cash = ce_r2.scalars().all()
    co10k_list = [e for e in all_cash
                  if e.entry_type.value == "cash_out" and float(e.amount) > 10000]
    co10k_cnt  = len(co10k_list)
    co10k_amt  = sum(float(e.amount) for e in co10k_list)

    # 3) SFT flagged customers
    sft_r2  = await db.execute(
        select(Customer).where(Customer.tenant_id == tid, Customer.sft_flagged == True))
    sft_cnt = len(sft_r2.scalars().all())
    sft_amt = sum(float(c.cash_receipts_fy) for c in sft_r2.scalars().all()) if False else None
    # re-fetch for amount
    sft_r3  = await db.execute(
        select(Customer).where(Customer.tenant_id == tid, Customer.sft_flagged == True))
    sft_custs2 = sft_r3.scalars().all()
    sft_amt    = sum(float(c.cash_receipts_fy) for c in sft_custs2)

    # 4) PAN missing on cash invoice >= 2L
    pan_miss_cnt = len([i for i in s269_invs if not (i.customer_pan or "").strip()])

    nc_data = [
        ("Section 269ST Violations",
         "Cash receipt >= Rs 2,00,000 — single transaction limit",
         s269_cnt, s269_amt),
        ("Cash Out > Rs 10,000",
         "Single cash-book outflow exceeding Rs 10,000 (Sec 269T-type)",
         co10k_cnt, co10k_amt),
        ("SFT Flagged Customers",
         "Sec 206 — cumulative cash receipts above SFT threshold",
         sft_cnt, sft_amt),
        ("PAN Missing on Large Cash Sale",
         "Cash invoice >= Rs 2L where customer PAN not recorded",
         pan_miss_cnt, None),
    ]

    for idx2, (area, desc, cnt, amt) in enumerate(nc_data):
        r    = 21 + idx2
        risk = cnt > 0
        rf   = F_RED_L if risk else F_WHITE
        ws_d.merge_cells(f"A{r}:C{r}")
        _D(r,1,area, font=Font(name="Calibri",bold=True,
           color="B71C1C" if risk else "388E3C",size=10), fill=rf, align=_lft)
        ws_d.merge_cells(f"D{r}:G{r}")
        _D(r,4,desc, font=Font(name="Calibri",bold=False,color="555555",size=9),
           fill=F_AMB_L if risk else F_LGREY, align=_lft)
        _D(r,8,cnt,  font=Font(name="Calibri",bold=True,
           color="B71C1C" if risk else "2E7D32",size=11), fill=rf, align=_ctr)
        ws_d.merge_cells(f"I{r}:J{r}")
        if amt is not None:
            _D(r,9,amt, font=Font(name="Calibri",bold=True,
               color="B71C1C" if amt>0 else "2E7D32",size=10),
               fill=rf, align=_rgt, fmt="#,##0.00")
        else:
            _D(r,9,"—", font=Font(name="Calibri",color="999999",size=10),
               fill=rf, align=_ctr)
        ws_d.row_dimensions[r].height = 20

    ws_d.row_dimensions[25].height = 6

    # ── Footer ───────────────────────────────────────────────
    ws_d.merge_cells("A26:J26")
    _D(26,1,f"GoldTrader Pro  ·  Generated {per}  ·  Taxly India | taxlyindia.com",
       font=Font(name="Calibri",bold=False,color="AAAAAA",size=8),
       fill=F_LGREY, align=_ctr)
    ws_d.row_dimensions[26].height = 14

    # Column widths
    for col, w in enumerate([18,14,14,16,12, 18,14,14,16,12], 1):
        ws_d.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 1: Sales Register ───────────────────────────────
    add_sheet(wb, "Sales Register", [
        "Invoice No", "Date", "Customer", "Mobile", "PAN", "Pay Mode",
        "Subtotal", "CGST", "SGST", "IGST", "Grand Total", "Status"
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name,
         inv.customer_mobile, inv.customer_pan or "", inv.pay_mode.value,
         float(inv.subtotal), float(inv.cgst), float(inv.sgst), float(inv.igst),
         float(inv.grand_total), inv.payment_status.value]
        for inv in invoices
    ])

    # ── Sheet 2: GSTR-1 ──────────────────────────────────────
    add_sheet(wb, "GSTR-1", [
        "Invoice No", "Date", "Customer", "GSTIN", "State", "HSN",
        "Taxable Value", "CGST%", "CGST", "SGST%", "SGST", "IGST", "Total"
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name,
         inv.customer_gstin or "Unregistered", inv.customer_state or "", "7113",
         float(inv.subtotal), float(inv.gst_rate/2), float(inv.cgst),
         float(inv.gst_rate/2), float(inv.sgst), float(inv.igst), float(inv.grand_total)]
        for inv in invoices
    ])

    # ── Sheet 3: Outstanding ──────────────────────────────────
    outstanding = [inv for inv in invoices if float(inv.outstanding) > 0]
    add_sheet(wb, "Outstanding", [
        "Invoice No", "Date", "Customer", "Mobile", "Grand Total", "Paid", "Outstanding"
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name,
         inv.customer_mobile, float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding)]
        for inv in outstanding
    ])

    # ── Sheet 4: Payments ─────────────────────────────────────
    pay_stmt = select(Payment).where(Payment.tenant_id == tid)
    if from_date: pay_stmt = pay_stmt.where(Payment.payment_date >= from_date)
    if to_date:   pay_stmt = pay_stmt.where(Payment.payment_date <= to_date)
    pay_result = await db.execute(pay_stmt.order_by(Payment.payment_date.desc()))
    payments   = pay_result.scalars().all()
    inv_no_map = {inv.id: inv.invoice_no for inv in invoices}

    pay_rows = []
    for p in payments:
        inv_obj = await db.get(Invoice, p.invoice_id) if p.invoice_id else None
        cname   = inv_obj.customer_name if inv_obj else "—"
        pay_rows.append([
            p.payment_date.isoformat(), inv_no_map.get(p.invoice_id,"—"),
            cname, p.customer_mobile, float(p.amount), p.pay_mode.value, p.reference_no or ""
        ])
    add_sheet(wb, "Payments", ["Date","Invoice No","Customer","Mobile","Amount","Mode","Reference"], pay_rows)

    # ── Sheet 5: Item-wise ────────────────────────────────────
    item_rows_data = []
    for inv in invoices:
        ir = await db.execute(select(InvoiceItem).where(InvoiceItem.invoice_id == inv.id))
        for item in ir.scalars():
            item_rows_data.append([
                inv.invoice_no, inv.invoice_date.isoformat(), inv.customer_name,
                item.category.value, item.purity or "", item.description,
                float(item.qty), item.unit.value, float(item.rate),
                float(item.making_charges), float(item.amount)
            ])
    add_sheet(wb, "Item-wise", [
        "Invoice No","Date","Customer","Category","Purity","Description",
        "Qty","Unit","Rate","Making","Amount"
    ], item_rows_data)

    # ── Sheet 6: SFT ─────────────────────────────────────────
    cust_result = await db.execute(
        select(Customer).where(Customer.tenant_id == tid, Customer.sft_flagged == True)
    )
    sft_custs = cust_result.scalars().all()
    add_sheet(wb, "SFT Register", [
        "Customer","Mobile","PAN","Cash Receipts FY","Threshold","PAN Missing"
    ], [
        [c.name, c.mobile, c.pan or "", float(c.cash_receipts_fy),
         float(SFT_THRESHOLD), "YES" if not c.pan else "No"]
        for c in sft_custs
    ])

    # ── Sheet 7: Section 269ST ────────────────────────────────
    threshold = Decimal("200000")
    viol_invs = [
        inv for inv in invoices
        if inv.pay_mode.value == "Cash" and float(inv.grand_total) >= float(threshold)
    ]
    viol_rows = []
    for inv in viol_invs:
        cust = await db.get(Customer, (inv.customer_mobile, tid))
        viol_rows.append([
            inv.invoice_date.isoformat(), inv.invoice_no,
            inv.customer_name, inv.customer_mobile,
            inv.customer_pan or (cust.pan if cust else "MISSING"),
            float(inv.grand_total), float(inv.grand_total), inv.notes or ""
        ])
    add_sheet(wb, "Sec 269ST Violations", [
        "Date","Invoice No","Customer","Mobile","PAN","Cash Amount","Penalty Risk","Notes"
    ], viol_rows)


    # ── Sheet 8: Cash Book ────────────────────────────────────
    cash_stmt = select(CashEntry).where(CashEntry.tenant_id == tid)
    if from_date: cash_stmt = cash_stmt.where(CashEntry.entry_date >= from_date)
    if to_date:   cash_stmt = cash_stmt.where(CashEntry.entry_date <= to_date)
    cash_result = await db.execute(cash_stmt.order_by(CashEntry.entry_date, CashEntry.id))
    cash_entries = cash_result.scalars().all()

    running = Decimal("0")
    cash_rows = []
    for e in cash_entries:
        amt   = Decimal(str(e.amount))
        etype = e.entry_type.value
        if etype in ("cash_in", "bank_in"):
            running += amt
        elif etype in ("cash_out", "cash_to_bank"):
            running -= amt
        cash_rows.append([
            e.entry_date.isoformat(), etype, e.description or "",
            float(amt) if etype in ("cash_in","bank_in") else 0,
            float(amt) if etype in ("cash_out","cash_to_bank") else 0,
            float(running), e.bank_reference or ""
        ])
    add_sheet(wb, "Cash Book", [
        "Date","Type","Description","Cash In","Cash Out","Balance","Reference"
    ], cash_rows)

    # ── Supplier Invoices ────────────────────────────────────
    sup_inv_r = await db.execute(
        select(SupplierInvoice)
        .where(SupplierInvoice.tenant_id == tid, SupplierInvoice.status == "active")
        .order_by(SupplierInvoice.invoice_date.desc())
    )
    sup_invs = sup_inv_r.scalars().all()
    add_sheet(wb, "Supplier Invoices", [
        "Invoice No", "Date", "Supplier Name", "Mobile",
        "Subtotal", "CGST", "SGST", "IGST", "Grand Total",
        "Amount Paid", "Outstanding", "Payment Status",
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.supplier_name, inv.supplier_mobile,
         float(inv.subtotal), float(inv.cgst), float(inv.sgst), float(inv.igst),
         float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding), inv.payment_status]
        for inv in sup_invs
    ])

    # ── Supplier Payments ─────────────────────────────────────
    sup_pay_r = await db.execute(
        select(SupplierPayment).where(SupplierPayment.tenant_id == tid)
        .order_by(SupplierPayment.payment_date.desc())
    )
    sup_pay_rows = []
    for p in sup_pay_r.scalars().all():
        sup_obj = await db.get(Supplier, (p.supplier_mobile, tid))
        sup_pay_rows.append([
            p.payment_date.isoformat(), sup_obj.name if sup_obj else "—", p.supplier_mobile,
            float(p.amount), p.pay_mode, p.reference_no or "—", p.notes or "",
        ])
    add_sheet(wb, "Supplier Payments", [
        "Date", "Supplier Name", "Mobile", "Amount", "Mode", "Reference", "Notes"
    ], sup_pay_rows)

    # ── Supplier Advances ─────────────────────────────────────
    sup_adv_r = await db.execute(
        select(SupplierAdvance).where(SupplierAdvance.tenant_id == tid)
        .order_by(SupplierAdvance.advance_date.desc())
    )
    sup_adv_rows = []
    for a in sup_adv_r.scalars().all():
        sup_obj = await db.get(Supplier, (a.supplier_mobile, tid))
        sup_adv_rows.append([
            a.advance_date.isoformat(), sup_obj.name if sup_obj else "—", a.supplier_mobile,
            float(a.amount), float(a.remaining), a.pay_mode, a.notes or "",
        ])
    add_sheet(wb, "Supplier Advances", [
        "Date", "Supplier Name", "Mobile", "Amount", "Remaining", "Mode", "Notes"
    ], sup_adv_rows)

    # ── Supplier Outstanding ──────────────────────────────────
    sup_out_r = await db.execute(
        select(SupplierInvoice)
        .where(SupplierInvoice.tenant_id == tid, SupplierInvoice.status == "active",
               SupplierInvoice.outstanding > 0)
        .order_by(SupplierInvoice.invoice_date)
    )
    add_sheet(wb, "Supplier Outstanding", [
        "Invoice No", "Invoice Date", "Supplier Name", "Mobile",
        "Grand Total", "Amount Paid", "Outstanding", "Days Overdue",
    ], [
        [inv.invoice_no, inv.invoice_date.isoformat(), inv.supplier_name, inv.supplier_mobile,
         float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding),
         (today_dt - inv.invoice_date).days]
        for inv in sup_out_r.scalars().all()
    ])

    # ── GSTR-2B Purchase Register ─────────────────────────────
    add_sheet(wb, "GSTR-2B Purchase", [
        "Invoice No", "Invoice Date", "Supplier Name", "Supplier GSTIN",
        "HSN Code", "Description", "Taxable Value",
        "CGST Rate%", "CGST Amt", "SGST Rate%", "SGST Amt",
        "IGST Rate%", "IGST Amt", "Invoice Total",
    ], [])
    ws_gstr2b = wb["GSTR-2B Purchase"]
    for inv in sup_invs:
        items_r = await db.execute(
            select(SupplierInvoiceItem).where(SupplierInvoiceItem.invoice_id == inv.id)
        )
        sup_obj   = await db.get(Supplier, (inv.supplier_mobile, tid))
        sup_gstin = sup_obj.gstin or "" if sup_obj else ""
        half_rate = float(inv.gst_rate) / 2
        for it in items_r.scalars().all():
            taxable  = float(it.amount)
            cgst_amt = round(taxable * half_rate / 100, 2) if inv.gst_type == "CGST+SGST" else 0
            sgst_amt = cgst_amt
            igst_amt = round(taxable * float(inv.gst_rate) / 100, 2) if inv.gst_type == "IGST" else 0
            ws_gstr2b.append([
                inv.invoice_no, inv.invoice_date.isoformat(),
                inv.supplier_name, sup_gstin,
                it.hsn_code, it.description, taxable,
                half_rate if inv.gst_type == "CGST+SGST" else 0, cgst_amt,
                half_rate if inv.gst_type == "CGST+SGST" else 0, sgst_amt,
                float(inv.gst_rate) if inv.gst_type == "IGST" else 0, igst_amt,
                float(inv.grand_total),
            ])
    auto_col_width(ws_gstr2b)

    # ── Customer Account Register sheet ─────────────────────
    ws_ca = wb.create_sheet("Customer Account")
    ca_headers = [
        "Invoice Date","Invoice No","Customer Name","Mobile",
        "Gold (Rs)","Silver (Rs)","Diamond (Rs)","Polish (Rs)",
        "Making (Rs)","CGST (Rs)","SGST (Rs)","IGST (Rs)","Grand Total (Rs)",
        "Paid (Rs)","Outstanding (Rs)"
    ]
    ws_ca.append(ca_headers)
    style_header_row(ws_ca, 1, len(ca_headers))
    ca_tot = {k: Decimal("0") for k in
              ["gold","silver","diamond","polish","making","cgst","sgst","igst","grand","paid","out"]}
    for inv in invoices:
        ir2 = await db.execute(select(InvoiceItem).where(InvoiceItem.invoice_id == inv.id))
        gold2=silver2=diamond2=polish2=making2 = Decimal("0")
        for item in ir2.scalars():
            base2 = (item.amount or Decimal("0")) - (item.making_charges or Decimal("0"))
            making2 += item.making_charges or Decimal("0")
            cat2 = item.category.value if hasattr(item.category,"value") else str(item.category)
            if   cat2 == "Gold":             gold2    += base2
            elif cat2 == "Silver":           silver2  += base2
            elif cat2 == "Diamond":          diamond2 += base2
            elif cat2 == "Polish Charges":   polish2  += base2
        ca_tot["gold"]   += gold2;   ca_tot["silver"]  += silver2
        ca_tot["diamond"]+= diamond2; ca_tot["polish"] += polish2
        ca_tot["making"] += making2;  ca_tot["cgst"]   += inv.cgst
        ca_tot["sgst"]   += inv.sgst; ca_tot["igst"]   += inv.igst
        ca_tot["grand"]  += inv.grand_total
        ca_tot["paid"]   += inv.amount_paid
        ca_tot["out"]    += inv.outstanding
        ws_ca.append([
            inv.invoice_date.isoformat(), inv.invoice_no,
            inv.customer_name, inv.customer_mobile,
            float(gold2), float(silver2), float(diamond2), float(polish2),
            float(making2), float(inv.cgst), float(inv.sgst), float(inv.igst),
            float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding),
        ])
    # Totals row
    ws_ca.append(["TOTAL","","","",
        float(ca_tot["gold"]),float(ca_tot["silver"]),float(ca_tot["diamond"]),float(ca_tot["polish"]),
        float(ca_tot["making"]),float(ca_tot["cgst"]),float(ca_tot["sgst"]),float(ca_tot["igst"]),
        float(ca_tot["grand"]),float(ca_tot["paid"]),float(ca_tot["out"])])
    tr_idx = ws_ca.max_row
    for col in range(1, len(ca_headers)+1):
        c2 = ws_ca.cell(tr_idx, col)
        c2.font = Font(name="Calibri", bold=True, size=10)
        c2.fill = PatternFill("solid", fgColor="FFF9C4")
    auto_col_width(ws_ca)

    # ── Supplier Account Register sheet ──────────────────────
    ws_sa = wb.create_sheet("Supplier Account")
    sa_headers = [
        "Invoice Date","Invoice No","Supplier Name","Mobile",
        "Gold (Rs)","Silver (Rs)","Diamond (Rs)","Polish (Rs)",
        "Making (Rs)","CGST (Rs)","SGST (Rs)","IGST (Rs)","Grand Total (Rs)",
        "Paid (Rs)","Outstanding (Rs)"
    ]
    ws_sa.append(sa_headers)
    style_header_row(ws_sa, 1, len(sa_headers))
    sa_tot = {k: Decimal("0") for k in
              ["gold","silver","diamond","polish","making","cgst","sgst","igst","grand","paid","out"]}
    for inv in sup_invs:
        ir3 = await db.execute(
            select(SupplierInvoiceItem).where(SupplierInvoiceItem.invoice_id == inv.id))
        gold3=silver3=diamond3=polish3=making3 = Decimal("0")
        for item in ir3.scalars():
            base3 = (item.amount or Decimal("0")) - (item.making_charges or Decimal("0"))
            making3 += item.making_charges or Decimal("0")
            cat3 = item.category.value if hasattr(item.category,"value") else str(item.category)
            if   cat3 == "Gold":             gold3    += base3
            elif cat3 == "Silver":           silver3  += base3
            elif cat3 == "Diamond":          diamond3 += base3
            elif cat3 == "Polish Charges":   polish3  += base3
        sa_tot["gold"]   += gold3;    sa_tot["silver"]  += silver3
        sa_tot["diamond"]+= diamond3; sa_tot["polish"]  += polish3
        sa_tot["making"] += making3;  sa_tot["cgst"]    += inv.cgst
        sa_tot["sgst"]   += inv.sgst; sa_tot["igst"]    += inv.igst
        sa_tot["grand"]  += inv.grand_total
        sa_tot["paid"]   += inv.amount_paid
        sa_tot["out"]    += inv.outstanding
        ws_sa.append([
            inv.invoice_date.isoformat(), inv.invoice_no,
            inv.supplier_name or "", inv.supplier_mobile,
            float(gold3), float(silver3), float(diamond3), float(polish3),
            float(making3), float(inv.cgst), float(inv.sgst), float(inv.igst),
            float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding),
        ])
    ws_sa.append(["TOTAL","","","",
        float(sa_tot["gold"]),float(sa_tot["silver"]),float(sa_tot["diamond"]),float(sa_tot["polish"]),
        float(sa_tot["making"]),float(sa_tot["cgst"]),float(sa_tot["sgst"]),float(sa_tot["igst"]),
        float(sa_tot["grand"]),float(sa_tot["paid"]),float(sa_tot["out"])])
    tr2_idx = ws_sa.max_row
    for col in range(1, len(sa_headers)+1):
        c3 = ws_sa.cell(tr2_idx, col)
        c3.font = Font(name="Calibri", bold=True, size=10)
        c3.fill = PatternFill("solid", fgColor="E8F5E9")
    auto_col_width(ws_sa)

    # ── Move Dashboard to first sheet position ────────────────
    if "Dashboard" in wb.sheetnames:
        wb.move_sheet("Dashboard", offset=-len(wb.sheetnames) + 1)

    filename = f"GoldTrader_All_Reports_{today_dt.isoformat()}.xlsx"
    return _stream_workbook(wb, filename)


# ── Account Register Excel ─────────────────────────────────────
# New endpoint for Account report Excel download

@router.get("/account-excel")
async def export_account_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """
    Export Account Register to Excel.
    One row per invoice: Invoice Date, Invoice No, Customer Name, Customer Mobile,
    Gold, Silver, Diamond, Polish Charges, Making Charges,
    CGST Amount, SGST Amount, IGST Amount, Grand Total.
    """
    from decimal import Decimal

    tenant_id = payload["tenant_id"]
    stmt = (
        select(Invoice)
        .where(Invoice.tenant_id == tenant_id, Invoice.status == "active")
        .order_by(Invoice.invoice_date.desc(), Invoice.id.desc())
    )
    if from_date:
        stmt = stmt.where(Invoice.invoice_date >= from_date)
    if to_date:
        stmt = stmt.where(Invoice.invoice_date <= to_date)

    result   = await db.execute(stmt)
    invoices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Account Register"

    headers = [
        "Invoice Date", "Invoice No", "Customer Name", "Customer Mobile",
        "Gold (₹)", "Silver (₹)", "Diamond (₹)", "Polish Charges (₹)",
        "Making Charges (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Grand Total (₹)"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # Totals accumulators
    tot = {k: Decimal("0") for k in
           ["gold","silver","diamond","polish","making","cgst","sgst","igst","grand"]}

    for inv in invoices:
        items_result = await db.execute(
            select(InvoiceItem).where(InvoiceItem.invoice_id == inv.id)
        )
        items = items_result.scalars().all()

        gold_amt = silver_amt = diamond_amt = polish_amt = making_total = Decimal("0")
        for item in items:
            cat = item.category.value
            item_base = item.amount - item.making_charges
            making_total += item.making_charges
            if cat == "Gold":           gold_amt    += item_base
            elif cat == "Silver":       silver_amt  += item_base
            elif cat == "Diamond":      diamond_amt += item_base
            elif cat == "Polish Charges": polish_amt += item_base

        tot["gold"]   += gold_amt;    tot["silver"]  += silver_amt
        tot["diamond"]+= diamond_amt; tot["polish"]  += polish_amt
        tot["making"] += making_total; tot["cgst"]   += inv.cgst
        tot["sgst"]   += inv.sgst;    tot["igst"]    += inv.igst
        tot["grand"]  += inv.grand_total

        ws.append([
            inv.invoice_date.isoformat(), inv.invoice_no,
            inv.customer_name, inv.customer_mobile,
            float(gold_amt), float(silver_amt), float(diamond_amt), float(polish_amt),
            float(making_total),
            float(inv.cgst), float(inv.sgst), float(inv.igst),
            float(inv.grand_total),
        ])

    # Totals row
    total_row = [
        "TOTAL", "", "", "",
        float(tot["gold"]), float(tot["silver"]), float(tot["diamond"]), float(tot["polish"]),
        float(tot["making"]),
        float(tot["cgst"]), float(tot["sgst"]), float(tot["igst"]),
        float(tot["grand"]),
    ]
    ws.append(total_row)
    # Bold the totals row
    total_row_idx = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="FFF0CC")  # light gold background

    auto_col_width(ws)
    date_range = f"{from_date or 'all'}_{to_date or 'all'}"
    # Add Dashboard sheet and move it before the Account Register
    await add_dashboard_sheet(wb, db, tenant_id, from_date, to_date)
    wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)
    return _stream_workbook(wb, f"account_register_{date_range}.xlsx")


# ── Supplier Account Excel ────────────────────────────────────

@router.get("/supplier-account-excel")
async def export_supplier_account_excel(
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    payload:   dict           = Depends(get_current_user_payload),
    db:        AsyncSession   = Depends(get_db),
):
    """Supplier Account Register Excel — mirrors Customer Account Register format."""
    from models import SupplierInvoice, SupplierInvoiceItem
    from decimal import Decimal

    tid      = payload["tenant_id"]
    today_dt = date.today()
    stmt     = (
        select(SupplierInvoice)
        .where(SupplierInvoice.tenant_id == tid, SupplierInvoice.status == "active")
        .order_by(SupplierInvoice.invoice_date.desc(), SupplierInvoice.id.desc())
    )
    if from_date: stmt = stmt.where(SupplierInvoice.invoice_date >= from_date)
    if to_date:   stmt = stmt.where(SupplierInvoice.invoice_date <= to_date)

    result   = await db.execute(stmt)
    invoices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Account"

    headers = [
        "Invoice Date", "Invoice No", "Supplier Name", "Supplier Mobile",
        "Gold (₹)", "Silver (₹)", "Diamond (₹)", "Polish Charges (₹)",
        "Making Charges (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)",
        "Grand Total (₹)", "Amount Paid (₹)", "Outstanding (₹)"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    tot = {k: Decimal("0") for k in
           ["gold","silver","diamond","polish","making","cgst","sgst","igst","grand","paid","out"]}

    for inv in invoices:
        items_r = await db.execute(select(SupplierInvoiceItem).where(SupplierInvoiceItem.invoice_id == inv.id))
        items   = items_r.scalars().all()
        gold = silver = diamond = polish = making = Decimal("0")
        for item in items:
            base   = (item.amount or Decimal("0")) - (item.making_charges or Decimal("0"))
            making += item.making_charges or Decimal("0")
            cat    = item.category.value if hasattr(item.category, "value") else str(item.category)
            if cat == "Gold":             gold    += base
            elif cat == "Silver":         silver  += base
            elif cat == "Diamond":        diamond += base
            elif cat == "Polish Charges": polish  += base

        cgst = inv.cgst or Decimal("0"); sgst = inv.sgst or Decimal("0"); igst = inv.igst or Decimal("0")
        for k, v in [("gold",gold),("silver",silver),("diamond",diamond),("polish",polish),
                     ("making",making),("cgst",cgst),("sgst",sgst),("igst",igst),
                     ("grand",inv.grand_total),("paid",inv.amount_paid),("out",inv.outstanding)]:
            tot[k] += v

        ws.append([
            inv.invoice_date.isoformat(), inv.invoice_no,
            inv.supplier_name or "", inv.supplier_mobile,
            float(gold), float(silver), float(diamond), float(polish),
            float(making), float(cgst), float(sgst), float(igst),
            float(inv.grand_total), float(inv.amount_paid), float(inv.outstanding),
        ])

    # Totals row
    total_row_idx = ws.max_row + 1
    ws.append([
        "TOTAL", "", "", "",
        float(tot["gold"]), float(tot["silver"]), float(tot["diamond"]), float(tot["polish"]),
        float(tot["making"]), float(tot["cgst"]), float(tot["sgst"]), float(tot["igst"]),
        float(tot["grand"]), float(tot["paid"]), float(tot["out"]),
    ])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.font  = Font(name="Calibri", bold=True, size=10)
        cell.fill  = PatternFill("solid", fgColor="FFF0CC")

    auto_col_width(ws)
    date_range = f"{from_date or 'all'}_{to_date or 'all'}"
    return _stream_workbook(wb, f"supplier_account_{date_range}.xlsx")
