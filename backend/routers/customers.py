# routers/customers.py — Customer CRUD (mobile = primary key per tenant)

from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

import pandas as pd

from database import get_db
from models import Customer, Invoice, Payment
from utils.auth import get_tenant_payload as get_current_user_payload
from utils.business import pan_is_mandatory, is_sft_flagged, current_fy

router = APIRouter()


# ── Schemas ───────────────────────────────────────────────────

class CustomerCreate(BaseModel):
    mobile:  str   = Field(..., pattern=r"^\d{10}$")   # PRIMARY KEY
    name:    str   = Field(..., min_length=1, max_length=200)
    state:   str   = Field(..., min_length=2)            # Mandatory for GST
    pan:     Optional[str] = Field(None, pattern=r"^[A-Z]{5}[0-9]{4}[A-Z]$")
    gstin:   Optional[str] = None
    address: Optional[str] = None
    email:   Optional[str] = None

class CustomerOut(BaseModel):
    mobile:           str
    name:             str
    pan:              Optional[str]
    state:            str
    gstin:            Optional[str]
    address:          Optional[str]
    cash_receipts_fy: float
    sft_flagged:      bool
    pan_mandatory:    bool

    class Config:
        from_attributes = True

    @classmethod
    def from_orm_with_extras(cls, c: Customer):
        return cls(
            mobile=c.mobile,
            name=c.name,
            pan=c.pan,
            state=c.state,
            gstin=c.gstin,
            address=c.address,
            cash_receipts_fy=float(c.cash_receipts_fy),
            sft_flagged=c.sft_flagged,
            pan_mandatory=pan_is_mandatory(c.cash_receipts_fy),
        )


# ── Create / Update Customer ──────────────────────────────────

@router.post("/", response_model=CustomerOut, status_code=201)
async def create_customer(
    body:    CustomerCreate,
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    """Create a new customer. Mobile is the unique identifier per tenant."""
    tenant_id = payload["tenant_id"]
    existing = await db.get(Customer, (body.mobile, tenant_id))
    if existing:
        raise HTTPException(status_code=409, detail="Customer with this mobile already exists.")

    customer = Customer(
        mobile=body.mobile,
        tenant_id=tenant_id,
        name=body.name,
        pan=body.pan,
        state=body.state,
        gstin=body.gstin,
        address=body.address,
        email=body.email,
        cash_receipts_fy=0,
        sft_flagged=False,
    )
    db.add(customer)
    await db.commit()
    await db.refresh(customer)
    return CustomerOut.from_orm_with_extras(customer)


@router.put("/{mobile}", response_model=CustomerOut)
async def update_customer(
    mobile:  str,
    body:    CustomerCreate,
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    tenant_id = payload["tenant_id"]
    customer  = await db.get(Customer, (mobile, tenant_id))
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    for field, value in body.dict(exclude_unset=True, exclude={"mobile"}).items():
        setattr(customer, field, value)

    await db.commit()
    await db.refresh(customer)
    return CustomerOut.from_orm_with_extras(customer)


# ── List Customers ────────────────────────────────────────────

@router.get("/", response_model=list[CustomerOut])
async def list_customers(
    q:       Optional[str] = None,
    payload: dict          = Depends(get_current_user_payload),
    db:      AsyncSession  = Depends(get_db),
):
    """List customers with optional name/mobile search."""
    tenant_id = payload["tenant_id"]
    stmt = select(Customer).where(Customer.tenant_id == tenant_id).order_by(Customer.name)

    if q:
        stmt = stmt.where(
            Customer.name.ilike(f"%{q}%") | Customer.mobile.contains(q)
        )

    result = await db.execute(stmt)
    customers = result.scalars().all()
    return [CustomerOut.from_orm_with_extras(c) for c in customers]


# ── Customer Ledger ───────────────────────────────────────────

@router.get("/{mobile}/ledger")
async def customer_ledger(
    mobile:  str,
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    """
    Return full transaction ledger for a customer:
    Invoices (debit) + Payments (credit) + Advances, sorted by date.
    """
    tenant_id = payload["tenant_id"]
    customer  = await db.get(Customer, (mobile, tenant_id))
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Invoices
    inv_result = await db.execute(
        select(Invoice).where(
            Invoice.tenant_id      == tenant_id,
            Invoice.customer_mobile == mobile,
        ).order_by(Invoice.invoice_date)
    )
    invoices = inv_result.scalars().all()

    # Payments
    pay_result = await db.execute(
        select(Payment).where(
            Payment.tenant_id      == tenant_id,
            Payment.customer_mobile == mobile,
        ).order_by(Payment.payment_date)
    )
    payments = pay_result.scalars().all()

    # Build ledger entries
    entries = []
    balance = 0.0

    for inv in invoices:
        balance += float(inv.grand_total)
        entries.append({
            "date":       inv.invoice_date.isoformat(),
            "type":       "Invoice",
            "ref":        inv.invoice_no,
            "debit":      float(inv.grand_total),
            "credit":     0,
            "balance":    balance,
        })

    for pay in payments:
        balance -= float(pay.amount)
        entries.append({
            "date":   pay.payment_date.isoformat(),
            "type":   "Payment",
            "ref":    f"PMT-{pay.id}",
            "debit":  0,
            "credit": float(pay.amount),
            "balance": balance,
        })

    entries.sort(key=lambda e: e["date"])

    return {
        "customer":        {"mobile": customer.mobile, "name": customer.name, "pan": customer.pan},
        "outstanding":     balance,
        "total_invoiced":  sum(e["debit"]  for e in entries),
        "total_paid":      sum(e["credit"] for e in entries),
        "entries":         entries,
    }


# ── Bulk Import via Excel ─────────────────────────────────────

REQUIRED_COLS = {"mobile", "name", "state"}

@router.post("/import-excel")
async def import_customers_excel(
    file:    UploadFile       = File(...),
    payload: dict             = Depends(get_current_user_payload),
    db:      AsyncSession     = Depends(get_db),
):
    """
    Bulk-import customers from an Excel file.
    Required columns: mobile, name, state
    Optional columns: pan, gstin, address, email
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Upload an Excel file (.xlsx or .xls)")

    contents = await file.read()
    df = pd.read_excel(BytesIO(contents), dtype=str)
    df.columns = [c.strip().lower() for c in df.columns]

    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(missing)}. "
                   f"Download the template from Customer Master → Import Excel."
        )

    tenant_id = payload["tenant_id"]
    created = 0
    updated = 0
    errors  = []

    for i, row in df.iterrows():
        mobile = str(row.get("mobile", "")).strip()
        name   = str(row.get("name",   "")).strip()
        state  = str(row.get("state",  "")).strip()

        if not mobile or not name or not state:
            errors.append(f"Row {i+2}: mobile, name, state are required")
            continue
        if len(mobile) != 10 or not mobile.isdigit():
            errors.append(f"Row {i+2}: invalid mobile '{mobile}'")
            continue

        existing = await db.get(Customer, (mobile, tenant_id))
        if existing:
            existing.name    = name
            existing.state   = state
            existing.pan     = str(row.get("pan", "")).strip() or existing.pan
            existing.gstin   = str(row.get("gstin", "")).strip() or existing.gstin
            existing.address = str(row.get("address", "")).strip() or existing.address
            updated += 1
        else:
            db.add(Customer(
                mobile=mobile, tenant_id=tenant_id,
                name=name, state=state,
                pan=str(row.get("pan", "")).strip() or None,
                gstin=str(row.get("gstin", "")).strip() or None,
                address=str(row.get("address", "")).strip() or None,
            ))
            created += 1

    await db.commit()
    return {"created": created, "updated": updated, "errors": errors}


# ── Customer Import Template Download ─────────────────────────

@router.get("/import-template")
async def download_import_template(
    payload: dict = Depends(get_current_user_payload),
):
    """
    Download a pre-formatted Excel template for bulk customer import.
    Contains headers, sample data row, and column instructions.
    """
    from fastapi.responses import Response as FastAPIResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Import"

    # Column definitions: (header, width, sample)
    columns = [
        ("mobile",  15, "9876543210"),
        ("name",    30, "Ramesh Shah"),
        ("state",   20, "Gujarat"),
        ("pan",     15, "ABCDE1234F"),
        ("gstin",   20, "24ABCDE1234F1Z5"),
        ("address", 40, "123 MG Road, Ahmedabad"),
        ("email",   30, "ramesh@example.com"),
    ]

    gold   = "C8900A"
    ltyel  = "FFF8E1"
    header_fill = PatternFill("solid", fgColor=gold)
    sample_fill = PatternFill("solid", fgColor=ltyel)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Row 1: Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "GoldTrader Pro — Customer Import Template"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1A1A2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: Instructions
    ws.merge_cells("A2:G2")
    instr = ws["A2"]
    instr.value = (
        "Instructions: Fill from Row 4 onwards. "
        "Required: mobile (10 digits), name, state. "
        "Optional: pan (PAN card), gstin, address, email. "
        "Do NOT change column headers."
    )
    instr.font = Font(italic=True, size=10, color="555555")
    instr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20

    # Row 3: Headers
    required_cols = {"mobile", "name", "state"}
    for col_idx, (header, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx)
        is_required = header in required_cols
        cell.value  = f"{header.upper()} {'*' if is_required else ''}".strip()
        cell.font   = Font(bold=True, size=11, color="FFFFFF")
        cell.fill   = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[3].height = 22

    # Row 4: Sample data
    for col_idx, (header, _, sample) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = sample
        cell.fill  = sample_fill
        cell.font  = Font(size=10, italic=True, color="555555")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[4].height = 18

    # Freeze header rows
    ws.freeze_panes = "A4"

    # Data validation note row
    ws.merge_cells("A5:G5")
    note = ws["A5"]
    note.value = "← Delete this sample row and rows 1-3 are locked. Enter your data from Row 4 onwards."
    note.font  = Font(size=9, color="999999", italic=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return FastAPIResponse(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="GoldTrader_Customer_Import_Template.xlsx"'
        },
    )
